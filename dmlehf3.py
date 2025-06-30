# -*- coding: utf-8 -*-
# ---------------------------------------------------------------------------
# 메이플스토리 수익 기록 프로그램
# 파일명: dmlehf.py (v0.9_consumable_enhancement) # 소모/획득 아이템 처리 개선 및 통계 반영
#
# 제작자: 의문의돌맹이
# 디스코드: gomsky.
#
# Copyright (c) 2025 의문의돌맹이. All rights reserved.
# 구조 변경 금지 / 상업적 재배포 금지 / 무단 수정 금지
# ---------------------------------------------------------------------------

import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox
import datetime
import openpyxl
import sys # sys 모듈 import 확인 (이미 있다면 생략)
import os
import re # MesoSaleForm 파싱 및 날짜/시간 포맷팅용
import json # 설정 파일 처리를 위해 추가

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # sys._MEIPASS is not defined, so running in development mode
        base_path = os.path.abspath(os.path.dirname(__file__)) # 개발 환경에서는 스크립트 파일 기준
    return os.path.join(base_path, relative_path)

APP_VERSION = "  영원히 탕후루 먹으면서 돌맹이 던지는 곰하늘 _ V1.2_UI_Guide" # 버전 업데이트 (UI 가이드 추가 명시)
DEFAULT_EXCEL_FILENAME = "메이플_수익_기록부_의문의돌맹이.xlsx"
CONFIG_FILE_NAME = "dmlehf_config.json"

# --- 시트 컬럼 정의 ---
HUNTING_SHEET_COLUMNS = [
    "날짜", "맵명", "시작시간", "종료시간", "시작메소", "종료메소", "판매후메소",
    "고가아이템", "고가아이템가치", "15분쿠폰사용횟수", "시작경험치", "종료경험치",
    "소모아이템비", "소모템획득수익", "지참비", "사냥메소수익", "일반템수익", "총수익", "순수익",
    "경험치수익", "원경험치"
]
JJUL_SHEET_COLUMNS = [
    "날짜", "맵명", "시작시간", "종료시간", "시작메소", "종료메소", "판매후메소",
    "쩔인원수", "1인당쩔비", "총쩔비", "고가아이템", "고가아이템가치",
    "소모아이템비", "소모템획득수익", "일반템수익", "총수익", "순수익"
]
RARE_ITEM_SHEET_COLUMNS = [
    "날짜", "세션유형", "맵명", "아이템명", "예상가치"
]
MESO_SALE_SHEET_COLUMNS = [
    "날짜", "100만메소당가격(원)", "판매량(단위: 100만메소)", "총판매액(원)"
]

# --- ExcelManager Class (엑셀 처리 담당) ---
class ExcelManager:
    def __init__(self, filepath):
        self.filename = filepath
        self.sheet_definitions = {
            "사냥세션": HUNTING_SHEET_COLUMNS,
            "쩔세션": JJUL_SHEET_COLUMNS,
            "고가템기록": RARE_ITEM_SHEET_COLUMNS,
            "메소판매기록": MESO_SALE_SHEET_COLUMNS,
            "일별요약": ["날짜", "사냥메소", "쩔수익", "고가템수익", "일반템수익",
                         "소모템획득수익", "소모아이템비", "지참비", "총수익", "순수익", "현금화금액(원)"],
            "맵별통계": ["맵명", "사냥횟수", "쩔횟수", "총사냥수익", "총쩔수익", "고가템수익",
                         "소모템획득수익", "평균사냥수익", "평균쩔수익"],
            "요일별통계": ["요일", "사냥수익", "쩔수익", "고가템수익", "일반템수익",
                           "소모템획득수익", "총수익", "순수익"],
            "ⓘ저작권 및 사용안내": ["제작자: 김하늘, 의문의돌맹이 (gomsky.)", "구조 변경 금지", "상업적 재배포 금지"]
        }

    def create_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            for sheet_name, columns in self.sheet_definitions.items():
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(columns)
                if sheet_name == "요일별통계":
                    weekdays = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
                    for day_name in weekdays:
                        initial_row_data = [day_name] + [0] * (len(columns) -1)
                        sheet.append(initial_row_data)

            workbook.save(self.filename)
            print(f"'{os.path.basename(self.filename)}' 템플릿이 자동 생성되었습니다 (ExcelManager - create_excel_template).")
            return True
        except PermissionError:
            print(f"DEBUG: PermissionError 발생 - 엑셀 파일 '{os.path.basename(self.filename)}' 생성/저장 불가 (ExcelManager - create_excel_template)")
            return False
        except Exception as e:
            print(f"DEBUG: Exception 발생 - 템플릿 자동 생성 중 알 수 없는 오류 (ExcelManager - create_excel_template): {e}")
            return False

    def save_data_to_sheet(self, data_to_save, sheet_name, columns_order, session_type=None, parent_for_msgbox=None):
        if not os.path.exists(self.filename):
            messagebox.showerror("파일 오류", f"엑셀 파일 '{self.filename}'을 찾을 수 없습니다.\n템플릿 파일을 먼저 생성해주세요.", parent=parent_for_msgbox)
            return False
        try:
            workbook = openpyxl.load_workbook(self.filename)
            if sheet_name not in workbook.sheetnames:
                 messagebox.showerror("엑셀 오류", f"엑셀 시트 '{sheet_name}'를 찾을 수 없습니다. 템플릿이 올바른지 확인해주세요.", parent=parent_for_msgbox)
                 return False
            current_sheet = workbook[sheet_name]

            row_data = []
            for col_name in columns_order:
                default_val = 0 if any(keyword in col_name for keyword in ["메소", "가치", "비", "수익", "인원수", "횟수", "경험치", "가격", "판매액", "판매량"]) else ""
                row_data.append(data_to_save.get(col_name, default_val))
            current_sheet.append(row_data)

            if session_type and 'rare_items_list' in data_to_save and data_to_save['rare_items_list']:
                if "고가템기록" not in workbook.sheetnames:
                    messagebox.showwarning("엑셀 경고", "'고가템기록' 시트를 찾을 수 없습니다. 고가 아이템 기록을 건너뜁니다.", parent=parent_for_msgbox)
                else:
                    sheet_rare = workbook["고가템기록"]
                    for item_name, item_value in data_to_save['rare_items_list']:
                        if item_name and isinstance(item_value, (int, float)) and item_value > 0:
                            rare_row_data = [
                                data_to_save.get('날짜', datetime.date.today().strftime("%Y-%m-%d")),
                                session_type,
                                data_to_save.get('맵명', "알수없음"),
                                item_name,
                                item_value
                            ]
                            sheet_rare.append(rare_row_data)
            elif session_type and data_to_save.get('고가아이템명') and data_to_save.get('고가아이템가치', 0) > 0 and \
                 not ('rare_items_list' in data_to_save and data_to_save['rare_items_list']):
                 if "고가템기록" not in workbook.sheetnames:
                    messagebox.showwarning("엑셀 경고", "'고가템기록' 시트를 찾을 수 없습니다. 고가 아이템 기록을 건너뜁니다.", parent=parent_for_msgbox)
                 else:
                    sheet_rare = workbook["고가템기록"]
                    rare_row_data = [
                        data_to_save.get('날짜'), session_type, data_to_save.get('맵명'),
                        data_to_save.get('고가아이템명'), data_to_save.get('고가아이템가치')
                    ]
                    sheet_rare.append(rare_row_data)

            workbook.save(self.filename)
            return True
        except PermissionError:
            print(f"DEBUG: PermissionError 발생 (save_data_to_sheet) - 파일 '{os.path.basename(self.filename)}' 접근 불가.")
            messagebox.showerror("엑셀 저장 오류", f"엑셀 파일 '{os.path.basename(self.filename)}'에 대한 쓰기 권한이 없거나 파일이 다른 프로그램에서 사용 중입니다.", parent=parent_for_msgbox)
            return False
        except Exception as e:
            messagebox.showerror("엑셀 저장 오류", f"데이터 저장 중 오류 발생: {e}", parent=parent_for_msgbox)
            import traceback
            traceback.print_exc()
            return False

    def update_daily_summary(self, date_str, hunting_meso=0, jjul_profit=0, rare_item_profit=0,
                             normal_item_profit=0, consumable_gained_profit=0,
                             consumable_cost=0, entry_fee=0, cash_sold_krw=0,
                             parent_for_msgbox=None):
        if not os.path.exists(self.filename):
            messagebox.showerror("파일 오류", f"엑셀 파일 '{self.filename}'을 찾을 수 없습니다.", parent=parent_for_msgbox)
            return False

        sheet_name = "일별요약"
        try:
            workbook = openpyxl.load_workbook(self.filename)
            if sheet_name not in workbook.sheetnames:
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트를 찾을 수 없습니다. 템플릿을 확인해주세요.", parent=parent_for_msgbox)
                return False

            summary_sheet = workbook[sheet_name]
            header = [cell.value for cell in summary_sheet[1]]

            required_cols = ["날짜", "사냥메소", "쩔수익", "고가템수익", "일반템수익",
                             "소모템획득수익", "소모아이템비", "지참비", "총수익", "순수익", "현금화금액(원)"]
            col_map = {name: i for i, name in enumerate(header)}
            if not all(k in col_map for k in required_cols):
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트의 헤더({required_cols})가 올바르지 않습니다.", parent=parent_for_msgbox)
                return False

            target_row_num = -1
            for row_num in range(2, summary_sheet.max_row + 1):
                cell_value = summary_sheet.cell(row=row_num, column=col_map["날짜"] + 1).value
                if isinstance(cell_value, datetime.datetime):
                    existing_date_str = cell_value.strftime("%Y-%m-%d")
                else:
                    existing_date_str = str(cell_value)

                if existing_date_str == date_str:
                    target_row_num = row_num
                    break

            if target_row_num != -1:
                def get_cell_value_as_int(row, col_key):
                    val_str = str(summary_sheet.cell(row=row, column=col_map[col_key] + 1).value)
                    try: return int(float(val_str)) if val_str and val_str != "None" else 0
                    except ValueError: return 0

                current_hunting_meso = get_cell_value_as_int(target_row_num, "사냥메소")
                current_jjul_profit = get_cell_value_as_int(target_row_num, "쩔수익")
                current_rare_item_profit = get_cell_value_as_int(target_row_num, "고가템수익")
                current_normal_item_profit = get_cell_value_as_int(target_row_num, "일반템수익")
                current_consumable_gained_profit = get_cell_value_as_int(target_row_num, "소모템획득수익")
                current_consumable_cost = get_cell_value_as_int(target_row_num, "소모아이템비")
                current_entry_fee = get_cell_value_as_int(target_row_num, "지참비")
                current_cash_sold_krw = get_cell_value_as_int(target_row_num, "현금화금액(원)")

                new_hunting_meso = current_hunting_meso + hunting_meso
                new_jjul_profit = current_jjul_profit + jjul_profit
                new_rare_item_profit = current_rare_item_profit + rare_item_profit
                new_normal_item_profit = current_normal_item_profit + normal_item_profit
                new_consumable_gained_profit = current_consumable_gained_profit + consumable_gained_profit
                new_consumable_cost = current_consumable_cost + consumable_cost
                new_entry_fee = current_entry_fee + entry_fee
                new_cash_sold_krw = current_cash_sold_krw + cash_sold_krw
            else:
                new_hunting_meso = hunting_meso
                new_jjul_profit = jjul_profit
                new_rare_item_profit = rare_item_profit
                new_normal_item_profit = normal_item_profit
                new_consumable_gained_profit = consumable_gained_profit
                new_consumable_cost = consumable_cost
                new_entry_fee = entry_fee
                new_cash_sold_krw = cash_sold_krw

            total_profit = (new_hunting_meso + new_jjul_profit + new_rare_item_profit +
                            new_normal_item_profit + new_consumable_gained_profit)
            net_profit = total_profit - new_consumable_cost - new_entry_fee

            row_values = [""] * len(header)
            row_values[col_map["날짜"]] = date_str
            row_values[col_map["사냥메소"]] = new_hunting_meso
            row_values[col_map["쩔수익"]] = new_jjul_profit
            row_values[col_map["고가템수익"]] = new_rare_item_profit
            row_values[col_map["일반템수익"]] = new_normal_item_profit
            row_values[col_map["소모템획득수익"]] = new_consumable_gained_profit
            row_values[col_map["소모아이템비"]] = new_consumable_cost
            row_values[col_map["지참비"]] = new_entry_fee
            row_values[col_map["총수익"]] = total_profit
            row_values[col_map["순수익"]] = net_profit
            row_values[col_map["현금화금액(원)"]] = new_cash_sold_krw

            if target_row_num != -1:
                for col_idx, value in enumerate(row_values):
                    summary_sheet.cell(row=target_row_num, column=col_idx + 1, value=value)
            else:
                summary_sheet.append(row_values)

            workbook.save(self.filename)
            print(f"'{sheet_name}' 시트가 '{date_str}' 날짜 기준으로 업데이트되었습니다.")
            return True
        except PermissionError:
            messagebox.showerror("일별 요약 업데이트 오류",
                                f"엑셀 파일 '{os.path.basename(self.filename)}'의 '일별요약' 시트를 업데이트할 수 없습니다.\n\n"
                                "파일이 다른 프로그램에서 열려있는지 확인해주세요.\n"
                                "열려있다면 해당 프로그램을 먼저 닫은 후, 다시 저장 버튼을 클릭해주시기 바랍니다.",
                                parent=parent_for_msgbox)
            return False
        except Exception as e:
            messagebox.showerror("일별 요약 업데이트 오류", f"'{sheet_name}' 시트 업데이트 중 오류 발생: {e}", parent=parent_for_msgbox)
            import traceback
            traceback.print_exc()
            return False

    def update_map_summary(self, map_name, session_type, profit_from_session=0,
                           rare_item_profit_from_session=0, consumable_gained_profit_from_session=0,
                           parent_for_msgbox=None):
        if not os.path.exists(self.filename):
            messagebox.showerror("파일 오류", f"엑셀 파일 '{self.filename}'을 찾을 수 없습니다.", parent=parent_for_msgbox)
            return False

        sheet_name = "맵별통계"
        try:
            workbook = openpyxl.load_workbook(self.filename)
            if sheet_name not in workbook.sheetnames:
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트를 찾을 수 없습니다. 템플릿을 확인해주세요.", parent=parent_for_msgbox)
                return False

            summary_sheet = workbook[sheet_name]
            header = [cell.value for cell in summary_sheet[1]]

            required_cols = ["맵명", "사냥횟수", "쩔횟수", "총사냥수익", "총쩔수익", "고가템수익",
                             "소모템획득수익", "평균사냥수익", "평균쩔수익"]
            col_map = {name: i for i, name in enumerate(header)}
            if not all(k in col_map for k in required_cols):
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트의 헤더({required_cols})가 올바르지 않습니다.", parent=parent_for_msgbox)
                return False

            target_row_num = -1
            for row_num in range(2, summary_sheet.max_row + 1):
                if str(summary_sheet.cell(row=row_num, column=col_map["맵명"] + 1).value) == map_name:
                    target_row_num = row_num
                    break

            if target_row_num != -1:
                def get_cell_val(row, col_key, data_type=int):
                    val_str = str(summary_sheet.cell(row=row, column=col_map[col_key] + 1).value)
                    try:
                        return data_type(float(val_str)) if val_str and val_str != "None" else (0 if data_type != str else "")
                    except ValueError: return (0 if data_type != str else "")

                current_hunt_count = get_cell_val(target_row_num, "사냥횟수")
                current_jjul_count = get_cell_val(target_row_num, "쩔횟수")
                current_total_hunt_profit = get_cell_val(target_row_num, "총사냥수익")
                current_total_jjul_profit = get_cell_val(target_row_num, "총쩔수익")
                current_total_rare_profit = get_cell_val(target_row_num, "고가템수익")
                current_total_consumable_gained_profit = get_cell_val(target_row_num, "소모템획득수익")
            else:
                current_hunt_count = 0
                current_jjul_count = 0
                current_total_hunt_profit = 0
                current_total_jjul_profit = 0
                current_total_rare_profit = 0
                current_total_consumable_gained_profit = 0

            new_hunt_count = current_hunt_count
            new_jjul_count = current_jjul_count
            new_total_hunt_profit = current_total_hunt_profit
            new_total_jjul_profit = current_total_jjul_profit
            new_total_rare_profit = current_total_rare_profit + rare_item_profit_from_session
            new_total_consumable_gained_profit = current_total_consumable_gained_profit + consumable_gained_profit_from_session

            if session_type == "사냥":
                new_hunt_count += 1
                new_total_hunt_profit += profit_from_session
            elif session_type == "쩔":
                new_jjul_count += 1
                new_total_jjul_profit += profit_from_session

            avg_hunt_profit = (new_total_hunt_profit + new_total_rare_profit + new_total_consumable_gained_profit) / new_hunt_count if new_hunt_count > 0 else 0
            avg_jjul_profit = (new_total_jjul_profit + new_total_rare_profit + new_total_consumable_gained_profit) / new_jjul_count if new_jjul_count > 0 else 0

            row_values = [""] * len(header)
            row_values[col_map["맵명"]] = map_name
            row_values[col_map["사냥횟수"]] = new_hunt_count
            row_values[col_map["쩔횟수"]] = new_jjul_count
            row_values[col_map["총사냥수익"]] = new_total_hunt_profit
            row_values[col_map["총쩔수익"]] = new_total_jjul_profit
            row_values[col_map["고가템수익"]] = new_total_rare_profit
            row_values[col_map["소모템획득수익"]] = new_total_consumable_gained_profit
            row_values[col_map["평균사냥수익"]] = int(avg_hunt_profit)
            row_values[col_map["평균쩔수익"]] = int(avg_jjul_profit)

            if target_row_num != -1:
                for col_idx, value in enumerate(row_values):
                    summary_sheet.cell(row=target_row_num, column=col_idx + 1, value=value)
            else:
                summary_sheet.append(row_values)

            workbook.save(self.filename)
            print(f"'{os.path.basename(self.filename)}' 파일의 '{sheet_name}' 시트가 '{map_name}' 맵 기준으로 업데이트되었습니다.")
            return True
        except PermissionError:
            messagebox.showerror("맵별 통계 업데이트 오류",
                                f"엑셀 파일 '{os.path.basename(self.filename)}'의 '맵별통계' 시트를 업데이트할 수 없습니다.\n\n"
                                "파일이 다른 프로그램에서 열려있는지 확인해주세요.\n"
                                "열려있다면 해당 프로그램을 먼저 닫은 후, 다시 저장 버튼을 클릭해주시기 바랍니다.",
                                parent=parent_for_msgbox)
            return False
        except Exception as e:
            messagebox.showerror("맵별 통계 업데이트 오류", f"'{sheet_name}' 시트 업데이트 중 오류 발생: {e}", parent=parent_for_msgbox)
            import traceback
            traceback.print_exc()
            return False

    def update_weekday_summary(self, date_str, session_type,
                               val_pure_revenue_contribution=0,
                               val_rare_item_contribution=0,
                               val_normal_item_contribution=0,
                               val_consumable_gained_contribution=0,
                               val_net_profit_contribution=0,
                               parent_for_msgbox=None):
        if not os.path.exists(self.filename):
            messagebox.showerror("파일 오류", f"엑셀 파일 '{self.filename}'을 찾을 수 없습니다.", parent=parent_for_msgbox)
            return False

        sheet_name = "요일별통계"
        try:
            workbook = openpyxl.load_workbook(self.filename)
            if sheet_name not in workbook.sheetnames:
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트를 찾을 수 없습니다. 템플릿을 확인해주세요.", parent=parent_for_msgbox)
                return False

            summary_sheet = workbook[sheet_name]
            header = [cell.value for cell in summary_sheet[1]]

            required_cols = ["요일", "사냥수익", "쩔수익", "고가템수익", "일반템수익",
                             "소모템획득수익", "총수익", "순수익"]
            col_map = {name: i for i, name in enumerate(header)}
            if not all(k in col_map for k in required_cols):
                messagebox.showerror("엑셀 오류", f"'{sheet_name}' 시트의 헤더({required_cols})가 올바르지 않습니다.", parent=parent_for_msgbox)
                return False

            try:
                dt_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                weekday_str = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"][dt_obj.weekday()]
            except ValueError:
                messagebox.showerror("오류", f"잘못된 날짜 형식입니다: {date_str}", parent=parent_for_msgbox)
                return False

            target_row_num = -1
            for row_num in range(2, summary_sheet.max_row + 1):
                if str(summary_sheet.cell(row=row_num, column=col_map["요일"] + 1).value) == weekday_str:
                    target_row_num = row_num
                    break
            
            if target_row_num != -1:
                def get_cell_val_as_int(row, col_key):
                    val_str = str(summary_sheet.cell(row=row, column=col_map[col_key] + 1).value)
                    try: return int(float(val_str)) if val_str and val_str != "None" else 0
                    except ValueError: return 0

                current_hunting_profit = get_cell_val_as_int(target_row_num, "사냥수익")
                current_jjul_profit = get_cell_val_as_int(target_row_num, "쩔수익")
                current_rare_item_profit = get_cell_val_as_int(target_row_num, "고가템수익")
                current_normal_item_profit = get_cell_val_as_int(target_row_num, "일반템수익")
                current_consumable_gained_profit = get_cell_val_as_int(target_row_num, "소모템획득수익")
                current_net_profit = get_cell_val_as_int(target_row_num, "순수익")
            else:
                current_hunting_profit = 0
                current_jjul_profit = 0
                current_rare_item_profit = 0
                current_normal_item_profit = 0
                current_consumable_gained_profit = 0
                current_net_profit = 0

            updated_hunting_profit = current_hunting_profit
            updated_jjul_profit = current_jjul_profit

            if session_type == "사냥":
                updated_hunting_profit += val_pure_revenue_contribution
            elif session_type == "쩔":
                updated_jjul_profit += val_pure_revenue_contribution

            updated_rare_item_profit = current_rare_item_profit + val_rare_item_contribution
            updated_normal_item_profit = current_normal_item_profit + val_normal_item_contribution
            updated_consumable_gained_profit = current_consumable_gained_profit + val_consumable_gained_contribution
            updated_net_profit = current_net_profit + val_net_profit_contribution

            updated_total_revenue = (updated_hunting_profit + updated_jjul_profit +
                                     updated_rare_item_profit + updated_normal_item_profit +
                                     updated_consumable_gained_profit)

            row_values = [""] * len(header)
            row_values[col_map["요일"]] = weekday_str
            row_values[col_map["사냥수익"]] = updated_hunting_profit
            row_values[col_map["쩔수익"]] = updated_jjul_profit
            row_values[col_map["고가템수익"]] = updated_rare_item_profit
            row_values[col_map["일반템수익"]] = updated_normal_item_profit
            row_values[col_map["소모템획득수익"]] = updated_consumable_gained_profit
            row_values[col_map["총수익"]] = updated_total_revenue
            row_values[col_map["순수익"]] = updated_net_profit

            if target_row_num != -1:
                for col_idx, value in enumerate(row_values):
                    summary_sheet.cell(row=target_row_num, column=col_idx + 1, value=value)
            else:
                summary_sheet.append(row_values)

            workbook.save(self.filename)
            print(f"'{os.path.basename(self.filename)}' 파일의 '{sheet_name}' 시트가 '{weekday_str}' 기준으로 업데이트되었습니다.")
            return True
        except PermissionError:
            messagebox.showerror("엑셀 저장 오류", f"엑셀 파일 '{self.filename}'에 대한 쓰기 권한이 없거나 파일이 다른 프로그램에서 사용 중입니다.", parent=parent_for_msgbox)
            return False
        except Exception as e:
            messagebox.showerror("요일별 통계 업데이트 오류", f"'{sheet_name}' 시트 업데이트 중 오류 발생: {e}", parent=parent_for_msgbox)
            import traceback
            traceback.print_exc()
            return False

# --- 스타일 변수 ---
DARK_BACKGROUND = '#2E2E2E'
DARK_FOREGROUND = '#FFFFFF'
DARK_ENTRY_BACKGROUND = '#3C3C3C'
DARK_BUTTON_BACKGROUND = '#4A4A4A'
DARK_BUTTON_ACTIVE_BACKGROUND = '#5A5A5A'
DARK_LISTBOX_BACKGROUND = '#3C3C3C'
DARK_LISTBOX_FOREGROUND = '#FFFFFF'
TOOLTIP_BG = "#FFFFE0"
TOOLTIP_FG = "#000000"
GUIDE_LABEL_FOREGROUND = "#AEDFF7" # 밝은 하늘색 계열로 가이드 라벨 텍스트 색상 정의

# --- Dynamic Entry List Frame 필드 설정 ---
RARE_ITEM_FIELDS_CONFIG = [
    {'label': '아이템명:', 'width': 25, 'entry_key': 'name_entry', 'default': '', 'tooltip': '획득한 고가 아이템의 이름', 'expand': True, 'fill': tk.X},
    {'label': '예상 가치:', 'width': 15, 'entry_key': 'value_entry', 'default': '0', 'tooltip': '해당 아이템의 예상 메소 가치'}
]
CONSUMABLE_ITEM_FIELDS_CONFIG = [
    {'label': '명칭:', 'width': 15, 'entry_key': 'name_entry', 'default': '', 
     'tooltip': '사용한 소모 아이템 또는 획득한 소모성 아이템의 이름', 'expand': True, 'fill': tk.X}, # 툴팁 수정
    {'label': '개당 가격/가치:', 'width': 12, 'entry_key': 'price_value_entry', 'default': '0', 
     'tooltip': "소모 시: 아이템 1개당 '구매 가격'\n획득 시: 아이템 1개당 '예상 판매 가치'"}, # 툴팁 수정
    {'label': '시작 수:', 'width': 7, 'entry_key': 'start_qty_entry', 'default': '0', 
     'tooltip': "해당 세션 '시작 시점'에 이 아이템을 몇 개 가지고 있었는지 입력하세요.\n(예: 사냥 전에 물약 100개 샀다면 '100' 입력)"}, # 툴팁 수정
    {'label': '종료 수:', 'width': 7, 'entry_key': 'end_qty_entry', 'default': '0', 
     'tooltip': "해당 세션 '종료 시점'에 이 아이템이 몇 개 남아있는지 입력하세요."} # 툴팁 수정
]

# --- Tooltip Class ---
class Tooltip:
    def __init__(self, widget, text, bg=TOOLTIP_BG, fg=TOOLTIP_FG):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.bg = bg
        self.fg = fg
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)
        widget.bind("<ButtonPress>", self.leave)

    def enter(self, event=None):
        if self.tooltip_window: return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background=self.bg, foreground=self.fg, relief='solid', borderwidth=1,
                         font=("맑은 고딕", "9", "normal"), padx=4, pady=2)
        label.pack(ipadx=1)

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

# --- Dynamic Entry List Frame Class (아이템 목록 관리 헬퍼) ---
class DynamicEntryListFrame(ttk.Frame):
    def __init__(self, parent_container, item_type_name_kor, fields_config_list,
                 no_item_text_format_str, canvas_for_scroll_update,
                 example_placeholder_str="", item_label_frame_parent=None):
        super().__init__(parent_container)
        self.item_type_name = item_type_name_kor
        self.fields_config = fields_config_list
        self.no_item_text_template = no_item_text_format_str
        self.canvas_to_update = canvas_for_scroll_update
        self.example_placeholder_text = example_placeholder_str
        self.item_label_frame_parent = item_label_frame_parent

        self.item_entries_widgets_list = []
        self.actual_list_frame = ttk.Frame(self)
        self.actual_list_frame.pack(fill=tk.X, expand=True)

        self.no_items_indicator_label = ttk.Label(
            self.actual_list_frame,
            text=self.no_item_text_template.format(self.item_type_name),
            style="Placeholder.TLabel"
        )
        self._toggle_no_items_label()

    def _toggle_no_items_label(self):
        if not self.item_entries_widgets_list:
            self.no_items_indicator_label.pack(pady=10, anchor=tk.W, padx=5)
        else:
            self.no_items_indicator_label.pack_forget()

        if self.canvas_to_update:
            self.canvas_to_update.after_idle(
                lambda: self.canvas_to_update.configure(scrollregion=self.canvas_to_update.bbox("all"))
            )

    def add_new_item_entry_row(self):
        entry_row_frame = ttk.Frame(self.actual_list_frame)
        entry_row_frame.pack(fill=tk.X, expand=True, pady=2)

        current_row_widgets = {"frame": entry_row_frame}
        for idx, field_conf in enumerate(self.fields_config):
            label_widget = ttk.Label(entry_row_frame, text=field_conf['label'])
            label_widget.pack(side=tk.LEFT, padx=(5 if idx > 0 else 0, 2))

            entry_widget_instance = ttk.Entry(entry_row_frame, width=field_conf['width'])
            entry_widget_instance.pack(side=tk.LEFT, padx=(0, 5),
                                       expand=field_conf.get('expand', False),
                                       fill=field_conf.get('fill', tk.NONE))

            if 'default' in field_conf:
                entry_widget_instance.insert(0, str(field_conf['default']))

            if field_conf['entry_key'] == 'name_entry' and \
               not self.item_entries_widgets_list and \
               self.example_placeholder_text:
                current_text = entry_widget_instance.get()
                if not current_text or current_text == str(field_conf.get('default', '')):
                    entry_widget_instance.delete(0, tk.END)
                    entry_widget_instance.insert(0, self.example_placeholder_text)

            Tooltip(entry_widget_instance, field_conf['tooltip'])
            current_row_widgets[field_conf['entry_key']] = entry_widget_instance

        remove_button_widget = ttk.Button(
            entry_row_frame, text=" X ", width=3,
            command=lambda frame=entry_row_frame: self._remove_specific_item_row(frame)
        )
        remove_button_widget.pack(side=tk.LEFT, padx=(5, 0))
        Tooltip(remove_button_widget, f"이 {self.item_type_name} 항목을 삭제합니다.")
        current_row_widgets["remove_button"] = remove_button_widget

        self.item_entries_widgets_list.append(current_row_widgets)
        self._toggle_no_items_label()

    def _remove_specific_item_row(self, frame_to_delete):
        for i, entry_widgets_set in enumerate(self.item_entries_widgets_list):
            if entry_widgets_set["frame"] == frame_to_delete:
                entry_widgets_set["frame"].destroy()
                del self.item_entries_widgets_list[i]
                break
        self._toggle_no_items_label()

    def get_all_entry_widgets(self):
        return list(self.item_entries_widgets_list)

    def clear_all_item_entries(self):
        for entry_widgets_set in list(self.item_entries_widgets_list):
            entry_widgets_set["frame"].destroy()
        self.item_entries_widgets_list.clear()
        self._toggle_no_items_label()

    def get_messagebox_parent(self):
        return self.item_label_frame_parent if self.item_label_frame_parent else self


class MainApplication:
    def __init__(self, root):
        self.root = root
        self.root.title(f"메이플 수익/판매 기록 by 의문의돌맹이 {APP_VERSION}") # 버전 표시 위치 변경
        self.root.geometry("850x900")

        self.excel_filepath_var = tk.StringVar()
        self.current_excel_filepath = ""

        self._load_config()
        
        global excel_manager
        excel_manager = ExcelManager(self.current_excel_filepath)
        self.excel_manager = excel_manager

        style = ttk.Style(self.root)
        azure_theme_applied_successfully = False

        try:
            theme_tcl_file_path = resource_path(os.path.join('theme', 'dark.tcl'))

            if not os.path.exists(theme_tcl_file_path):
                print(f"테마 파일 '{theme_tcl_file_path}'을(를) 찾을 수 없습니다.")
            else:
                self.root.tk.call('source', theme_tcl_file_path)
            
            theme_name_to_use = 'azure-dark' 
            available_themes = style.theme_names()

            if theme_name_to_use in available_themes:
                style.theme_use(theme_name_to_use)
                azure_theme_applied_successfully = True
            elif 'forest-dark' in available_themes: 
                style.theme_use('forest-dark')
                azure_theme_applied_successfully = True
                print(f"'{theme_name_to_use}' 테마를 찾을 수 없어 'forest-dark' 테마로 대체 적용합니다.")
            
            style.configure('.', background=DARK_BACKGROUND, foreground=DARK_FOREGROUND, font=('맑은 고딕', 9))
            style.configure('TFrame', background=DARK_BACKGROUND)
            style.configure('TLabel', background=DARK_BACKGROUND, foreground=DARK_FOREGROUND)
            style.configure('TButton', background=DARK_BUTTON_BACKGROUND, foreground=DARK_FOREGROUND, borderwidth=1)
            style.map('TButton', background=[('active', DARK_BUTTON_ACTIVE_BACKGROUND)])
            style.configure('TEntry', fieldbackground=DARK_ENTRY_BACKGROUND, foreground=DARK_FOREGROUND, insertcolor=DARK_FOREGROUND, borderwidth=1)
            style.configure('TNotebook', background=DARK_BACKGROUND, tabmargins=[2, 5, 2, 0])
            style.configure('TNotebook.Tab', background=DARK_BUTTON_BACKGROUND, foreground=DARK_FOREGROUND, padding=[10,5], font=('맑은 고딕', 10, 'bold'))
            style.map('TNotebook.Tab', background=[('selected', DARK_BUTTON_ACTIVE_BACKGROUND), ('active', '#5A5A5A')])
            style.configure('Treeview', background=DARK_LISTBOX_BACKGROUND, fieldbackground=DARK_LISTBOX_BACKGROUND, foreground=DARK_LISTBOX_FOREGROUND)
            style.configure('Treeview.Heading', background=DARK_BUTTON_BACKGROUND, foreground=DARK_FOREGROUND, font=('맑은 고딕', 10, 'bold'))
            style.configure('Vertical.TScrollbar', background=DARK_BUTTON_BACKGROUND, troughcolor=DARK_BACKGROUND, bordercolor=DARK_ENTRY_BACKGROUND, arrowcolor=DARK_FOREGROUND)
            style.configure('Horizontal.TScrollbar', background=DARK_BUTTON_BACKGROUND, troughcolor=DARK_BACKGROUND, bordercolor=DARK_ENTRY_BACKGROUND, arrowcolor=DARK_FOREGROUND)
            style.configure('TLabelframe', background=DARK_BACKGROUND, bordercolor=DARK_BUTTON_ACTIVE_BACKGROUND)
            style.configure('TLabelframe.Label', background=DARK_BACKGROUND, foreground=DARK_FOREGROUND, font=('맑은 고딕', 10, 'bold'))
            style.configure("Placeholder.TLabel", foreground="gray", background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'italic'))
            style.configure("Error.TLabel", foreground="red", background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'bold'))
            # --- 가이드 라벨 스타일 추가 ---
            style.configure("Guide.TLabel", foreground=GUIDE_LABEL_FOREGROUND, background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'italic'))
            
            if azure_theme_applied_successfully:
                print(f"TTK 테마 적용 및 추가 스타일 구성 완료.")
            elif os.path.exists(theme_tcl_file_path):
                 print(f"'{theme_name_to_use}' 테마를 적용할 수 없습니다. 수동 다크모드 스타일만 적용됩니다.")
            else:
                 print(f"테마 파일을 찾을 수 없어 수동 다크모드 스타일만 적용됩니다.")

        except tk.TclError as e:
            print(f"TTK 테마 로드 또는 적용 중 오류 발생: {e}. 수동 다크모드 스타일을 적용합니다.")
            style.configure('.', background=DARK_BACKGROUND, foreground=DARK_FOREGROUND, font=('맑은 고딕', 9))
            style.configure('TFrame', background=DARK_BACKGROUND)
            style.configure('TLabel', background=DARK_BACKGROUND, foreground=DARK_FOREGROUND)
            style.configure('TButton', background=DARK_BUTTON_BACKGROUND, foreground=DARK_FOREGROUND, borderwidth=1, relief='solid')
            style.map('TButton', background=[('active', DARK_BUTTON_ACTIVE_BACKGROUND)])
            style.configure('TEntry', fieldbackground=DARK_ENTRY_BACKGROUND, foreground=DARK_FOREGROUND, insertcolor=DARK_FOREGROUND, borderwidth=1, relief='solid')
            style.configure('TNotebook', background=DARK_BACKGROUND, tabmargins=[2, 5, 2, 0])
            style.configure('TNotebook.Tab', background=DARK_BUTTON_BACKGROUND, foreground=DARK_FOREGROUND, padding=[10,5], font=('맑은 고딕', 10, 'bold'))
            style.map('TNotebook.Tab', background=[('selected', DARK_BUTTON_ACTIVE_BACKGROUND), ('active', '#5A5A5A')])
            style.configure("Placeholder.TLabel", foreground="gray", background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'italic'))
            style.configure("Error.TLabel", foreground="red", background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'bold'))
            style.configure("Guide.TLabel", foreground=GUIDE_LABEL_FOREGROUND, background=DARK_BACKGROUND, font=('맑은 고딕', 9, 'italic'))


        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)

        self.hunt_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.hunt_tab_frame, text='사냥 결과 입력')
        self.hunting_form = HuntingResultForm(self.hunt_tab_frame, self)

        self.jjul_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.jjul_tab_frame, text='쩔 결과 입력')
        self.jjul_form = JjulResultForm(self.jjul_tab_frame, self)

        self.meso_sale_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.meso_sale_tab_frame, text='메소 판매 기록')
        self.meso_sale_form = MesoSaleForm(self.meso_sale_tab_frame, self)

        self.usage_guide_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.usage_guide_tab_frame, text='📖 사용 방법')
        usage_text_frame = ttk.Frame(self.usage_guide_tab_frame, padding=(10,10))
        usage_text_frame.pack(expand=True, fill=tk.BOTH)
        self.usage_guide_text = tk.Text(
            usage_text_frame, wrap=tk.WORD, padx=15, pady=15, font=("맑은 고딕", 10),
            relief=tk.FLAT, bg=DARK_BACKGROUND, fg=DARK_FOREGROUND, exportselection=False,
            spacing1=5, spacing2=3, spacing3=10, insertbackground=DARK_FOREGROUND
        )
        usage_scrollbar = ttk.Scrollbar(usage_text_frame, orient="vertical", command=self.usage_guide_text.yview)
        self.usage_guide_text.configure(yscrollcommand=usage_scrollbar.set)
        usage_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.usage_guide_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._populate_usage_guide_tab_content() # 초기화 시 호출
        self.usage_guide_text.config(state=tk.DISABLED)

        self.program_info_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.program_info_tab_frame, text='ⓘ 제작자/정보')
        info_text_frame = ttk.Frame(self.program_info_tab_frame, padding=(10,10))
        info_text_frame.pack(expand=True, fill=tk.BOTH)
        self.program_info_text = tk.Text(
            info_text_frame, wrap=tk.WORD, padx=15, pady=15, font=("맑은 고딕", 10),
            relief=tk.FLAT, bg=DARK_BACKGROUND, fg=DARK_FOREGROUND, exportselection=False,
            spacing1=5, spacing2=3, spacing3=10, insertbackground=DARK_FOREGROUND
        )
        info_scrollbar = ttk.Scrollbar(info_text_frame, orient="vertical", command=self.program_info_text.yview)
        self.program_info_text.configure(yscrollcommand=info_scrollbar.set)
        info_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.program_info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._populate_program_info_tab_content()
        self.program_info_text.config(state=tk.DISABLED)

        self.settings_tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_tab_frame, text='⚙️ 설정')
        settings_content_frame = ttk.Frame(self.settings_tab_frame, padding=(20, 20))
        settings_content_frame.pack(expand=True, fill=tk.BOTH)
        settings_content_frame.columnconfigure(1, weight=1)
        row_idx = 0
        excel_path_label = ttk.Label(settings_content_frame, text="엑셀 파일 저장 경로:")
        excel_path_label.grid(row=row_idx, column=0, padx=(0, 10), pady=10, sticky=tk.W)
        self.excel_path_entry = ttk.Entry(settings_content_frame, textvariable=self.excel_filepath_var, state="readonly", width=60)
        self.excel_path_entry.grid(row=row_idx, column=1, padx=5, pady=10, sticky=tk.EW)
        Tooltip(self.excel_path_entry, "현재 설정된 엑셀 파일의 전체 경로입니다. 아래 버튼으로 변경할 수 있습니다.")
        change_path_button = ttk.Button(settings_content_frame, text="경로 변경 및 파일명 설정", command=self._change_excel_path, width=25)
        change_path_button.grid(row=row_idx, column=2, padx=(10, 0), pady=10, sticky=tk.E)
        Tooltip(change_path_button, "엑셀 파일을 저장할 폴더와 파일 이름을 새로 설정합니다.")
        row_idx += 1
        self.settings_status_label = ttk.Label(settings_content_frame, text="", style="Placeholder.TLabel")
        self.settings_status_label.grid(row=row_idx, column=0, columnspan=3, pady=(15, 0), sticky=tk.W)
        self._update_settings_status_label("설정이 로드되었습니다.")

    @staticmethod
    def static_get_entry_value(entry_widget, data_type=str, default_value=None, is_time=False, field_name_for_error="", parent_for_msgbox=None):
        value = entry_widget.get().strip()
        label_text = field_name_for_error
        if not value:
            if default_value is not None:
                if data_type == int: return int(default_value)
                if data_type == float: return float(default_value)
                return default_value
            elif is_time:
                return ""
            else:
                return "" if data_type == str else (0 if data_type in [int, float] else None)

        placeholders = ["예: 혼돈의 파편", "예: 파워엘릭서", "예: 뒤틀린 낙인의 영혼석", "예: 만병통치약"]
        if entry_widget.winfo_class() == "TEntry" and value in placeholders:
            return "" if data_type == str else (0 if data_type in [int, float] else None)

        if is_time:
            try:
                datetime.datetime.strptime(value, "%H:%M").time()
                return value
            except ValueError:
                messagebox.showerror("입력 오류", f"'{label_text or '시간'}' 항목에 유효한 시간 형식을 입력해주세요 (HH:MM).", parent=parent_for_msgbox)
                raise ValueError(f"잘못된 시간 형식: {label_text or '시간'}")
        try:
            if data_type == int: return int(value.replace(",", ""))
            elif data_type == float: return float(value.replace(",", ""))
            return value
        except ValueError:
            messagebox.showerror("입력 오류", f"'{label_text or '숫자'}' 항목에 유효한 {data_type.__name__} 값을 입력해주세요.", parent=parent_for_msgbox)
            raise ValueError(f"잘못된 {data_type.__name__} 값: {label_text or '숫자'}")

    @staticmethod
    def static_format_date_entry(event_widget, entry_widget_to_format, parent_for_msgbox=None):
        current_text = entry_widget_to_format.get().strip()
        original_text_for_error = current_text
        
        if not current_text:
            entry_widget_to_format.delete(0, tk.END)
            entry_widget_to_format.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
            return

        formatted_date = ""
        current_year = str(datetime.date.today().year)
        
        patterns_with_year = [
            (r"(\d{4})-(\d{1,2})-(\d{1,2})", lambda m: datetime.datetime(int(m[0]), int(m[1]), int(m[2]))),
            (r"(\d{4})(\d{2})(\d{2})", lambda m: datetime.datetime(int(m[0]), int(m[1]), int(m[2]))),
            (r"(\d{4})\.(\d{1,2})\.(\d{1,2})", lambda m: datetime.datetime(int(m[0]), int(m[1]), int(m[2]))),
        ]
        patterns_without_year = [
            (r"(\d{2})-(\d{1,2})-(\d{1,2})", lambda m: datetime.datetime(int(current_year[:2] + m[0]), int(m[1]), int(m[2]))),
            (r"(\d{2})(\d{2})(\d{2})", lambda m: datetime.datetime(int(current_year[:2] + m[0]), int(m[1]), int(m[2]))),
            (r"(\d{2})\.(\d{1,2})\.(\d{1,2})", lambda m: datetime.datetime(int(current_year[:2] + m[0]), int(m[1]), int(m[2]))),
            (r"(\d{1,2})-(\d{1,2})", lambda m: datetime.datetime(int(current_year), int(m[0]), int(m[1]))),
            (r"(\d{1,2})\.(\d{1,2})", lambda m: datetime.datetime(int(current_year), int(m[0]), int(m[1]))),
            (r"(\d{2})(\d{2})", lambda m: datetime.datetime(int(current_year), int(m[0]), int(m[1]))),
            (r"(\d{1})(\d{2})", lambda m: datetime.datetime(int(current_year), int(m[0]), int(m[1]))),
        ]

        all_patterns = patterns_with_year + patterns_without_year
        date_obj = None

        for pattern_str, handler in all_patterns:
            match = re.fullmatch(pattern_str, current_text)
            if match:
                try:
                    groups = list(match.groups())
                    if pattern_str in [r"(\d{1,2})-(\d{1,2})", r"(\d{1,2})\.(\d{1,2})", r"(\d{2})(\d{2})", r"(\d{1})(\d{2})"]:
                        if len(groups[0]) == 1 and pattern_str != r"(\d{1})(\d{2})": groups[0] = groups[0].zfill(2)
                        if len(groups[1]) == 1 : groups[1] = groups[1].zfill(2)
                    if pattern_str == r"(\d{1})(\d{2})":
                        groups[0] = groups[0].zfill(2)
                    date_obj = handler(groups)
                    formatted_date = date_obj.strftime("%Y-%m-%d")
                    break 
                except ValueError:
                    continue 
        
        if not formatted_date:
            formatted_date = datetime.date.today().strftime("%Y-%m-%d")

        entry_widget_to_format.delete(0, tk.END)
        entry_widget_to_format.insert(0, formatted_date)

    @staticmethod
    def static_format_time_entry(event_widget, entry_widget_to_format, parent_for_msgbox=None):
        current_text = entry_widget_to_format.get().strip()
        original_text_for_error = current_text

        if not current_text:
            return

        formatted_time = ""
        
        if re.fullmatch(r"([01]\d|2[0-3]):([0-5]\d)", current_text):
            formatted_time = current_text
        elif re.fullmatch(r"([01]\d|2[0-3])([0-5]\d)", current_text) and len(current_text) == 4:
            m = re.match(r"([01]\d|2[0-3])([0-5]\d)", current_text)
            formatted_time = f"{m.group(1)}:{m.group(2)}"
        elif re.fullmatch(r"(\d)([0-5]\d)", current_text) and len(current_text) == 3:
            m = re.match(r"(\d)([0-5]\d)", current_text)
            formatted_time = f"0{m.group(1)}:{m.group(2)}"
        else:
            pass 

        if formatted_time:
            try: 
                datetime.datetime.strptime(formatted_time, "%H:%M")
                entry_widget_to_format.delete(0, tk.END)
                entry_widget_to_format.insert(0, formatted_time)
            except ValueError:
                 pass
            
    def _populate_usage_guide_tab_content(self): # 이 함수는 Phase 2에서 내용 대폭 강화 예정
        self.usage_guide_text.tag_configure("main_title", font=("맑은 고딕", 18, "bold"), foreground="#64B5F6", spacing3=20, justify=tk.CENTER)
        self.usage_guide_text.tag_configure("step_title", font=("맑은 고딕", 11, "bold"), foreground="#AED581", spacing1=10, spacing3=6)
        self.usage_guide_text.tag_configure("item", lmargin1=15, lmargin2=30, font=("맑은 고딕", 10), spacing2=4)
        self.usage_guide_text.tag_configure("tip_marker", font=("맑은 고딕", 10, "bold"), foreground="#FFB74D")
        self.usage_guide_text.tag_configure("tip_text", lmargin1=30, lmargin2=45, font=("맑은 고딕", 9, "italic"), foreground="#B0BEC5", spacing2=3)
        self.usage_guide_text.tag_configure("highlight", font=("맑은 고딕", 10, "bold"), foreground="#FFD54F")
        self.usage_guide_text.tag_configure("warning_marker", font=("맑은 고딕", 10, "bold"), foreground="#FF7043")
        self.usage_guide_text.tag_configure("warning_text", lmargin1=30, lmargin2=45, font=("맑은 고딕", 9, "bold", "italic"), foreground="#D32F2F", spacing2=3)

        # --- 사용 방법 탭 내용 (Phase 1에서는 기존 내용 유지 또는 최소한의 업데이트) ---
        # Phase 2에서 이 부분을 시나리오 기반으로 전면 개편할 예정입니다.
        # 현재는 기존 내용과 유사하게 유지하거나, Phase 1의 핵심 가이드라인을 간단히 언급하는 수준으로 둘 수 있습니다.
        # 여기서는 기존 내용 골자를 유지하되, 파일명 업데이트 로직은 _change_excel_path에서 처리하도록 합니다.
        
        # 기존 내용 (요약)
        self.usage_guide_text.delete('1.0', tk.END) # 기존 내용 삭제 후 다시 채우기
        self.usage_guide_text.insert(tk.END, "프로그램 사용 방법 (간단 안내)\n", "main_title")
        self.usage_guide_text.insert(tk.END, "자세한 사용법과 예시는 향후 업데이트될 예정입니다.\n\n", "item")
        
        self.usage_guide_text.insert(tk.END, "1. 핵심 입력 원칙\n", "step_title")
        self.usage_guide_text.insert(tk.END,
            "   • ", "item")
        self.usage_guide_text.insert(tk.END, "시작 메소", "highlight")
        self.usage_guide_text.insert(tk.END, ": 모든 준비(물약, 지참비 등)를 마친 후, 세션 시작 시 실제 보유 메소.\n", "item")
        self.usage_guide_text.insert(tk.END,
            "   • ", "item")
        self.usage_guide_text.insert(tk.END, "판매 후 메소", "highlight")
        self.usage_guide_text.insert(tk.END, ": 세션 중 얻은 일반 아이템(잡템) 판매 후 최종 보유 메소.\n", "item")
        self.usage_guide_text.insert(tk.END,
            "   • 각 입력칸에 마우스를 올리면 상세 설명(툴팁)이 나타납니다.\n", "item")

        excel_file_basename = os.path.basename(self.current_excel_filepath) if self.current_excel_filepath else DEFAULT_EXCEL_FILENAME
        self.usage_guide_text.insert(tk.END, f"\n2. 저장 및 통계\n", "step_title")
        self.usage_guide_text.insert(tk.END,
            f"   • '저장 및 계산' 버튼으로 '{excel_file_basename}' 파일에 기록하고 통계 시트가 자동 업데이트됩니다.\n", "item")

        self.usage_guide_text.insert(tk.END, "\n3. 문제 발생 시\n", "step_title")
        self.usage_guide_text.insert(tk.END,
            f"   • 엑셀 파일('{excel_file_basename}')이 열려있으면 저장 오류가 발생할 수 있습니다. 닫고 시도해주세요.\n", "item")
        # --- 여기까지 임시 내용 ---

    def _populate_program_info_tab_content(self):
        self.program_info_text.tag_configure("main_title", font=("맑은 고딕", 18, "bold"), foreground="#64B5F6", spacing3=20, justify=tk.CENTER)
        self.program_info_text.tag_configure("section_title", font=("맑은 고딕", 13, "bold"), foreground="#81C784", spacing1=18, spacing3=12)
        self.program_info_text.tag_configure("header", font=("맑은 고딕", 11, "bold"), foreground="#FFAB91", spacing1=10, spacing3=6)
        self.program_info_text.tag_configure("item", lmargin1=15, lmargin2=30, font=("맑은 고딕", 10), spacing2=4)
        self.program_info_text.tag_configure("highlight", font=("맑은 고딕", 10, "bold"), foreground="#FFD54F")
        self.program_info_text.tag_configure("donation_header", font=("맑은 고딕", 11, "bold"), foreground="#4DD0E1", spacing1=15, spacing3=6)
        self.program_info_text.tag_configure("donation_item", lmargin1=15, lmargin2=30, font=("맑은 고딕", 10), spacing2=4)
        self.program_info_text.tag_configure("donation_account", font=("맑은 고딕", 10, "bold"), foreground="#FFD54F", spacing2=4)
        self.program_info_text.tag_configure("copyright_item", lmargin1=15, lmargin2=30, font=("맑은 고딕", 9), spacing2=3)
        self.program_info_text.tag_configure("copyright_notice", font=("맑은 고딕", 9, "bold"), foreground="#EF9A9A", spacing1=5)
        self.program_info_text.tag_configure("final_greeting", font=("맑은 고딕", 10, "italic"), foreground="#90A4AE", justify=tk.CENTER, spacing1=25, spacing3=10)

        self.program_info_text.insert(tk.END, "프로그램 및 제작자 정보\n", "main_title")
        self.program_info_text.insert(tk.END, "ℹ️ 프로그램 정보\n", "section_title")
        self.program_info_text.insert(tk.END, "프로그램명: ", "header")
        self.program_info_text.insert(tk.END, "메이플 수익/판매 기록 프로그램\n", "item")
        self.program_info_text.insert(tk.END, "버      전: ", "header")
        self.program_info_text.insert(tk.END, f"{APP_VERSION}\n", "item")
        self.program_info_text.insert(tk.END, "👨‍💻 제작자\n", "section_title")
        self.program_info_text.insert(tk.END, "닉네임: ", "header")
        self.program_info_text.insert(tk.END, "의문의돌맹이\n", ("item", "highlight"))
        self.program_info_text.insert(tk.END, "이   름: ", "header")
        self.program_info_text.insert(tk.END, "김하늘\n", ("item", "highlight"))
        self.program_info_text.insert(tk.END, "Discord: ", "header")
        self.program_info_text.insert(tk.END, "gomsky.\n", ("item", "highlight"))
        self.program_info_text.insert(tk.END, "💖 제작자 후원 안내\n", "section_title")
        self.program_info_text.insert(tk.END,
            "프로그램이 마음에 드셨다면, 커피 한 잔의 따뜻한 후원으로 개발자에게 힘을 실어주실 수 있습니다.\n"
            "보내주신 후원금은 더 나은 프로그램을 만드는 데 소중히 사용하겠습니다. 감사합니다!\n\n",
            "donation_item"
        )
        self.program_info_text.insert(tk.END, "후원 계좌: ", ("donation_item", "donation_header"))
        self.program_info_text.insert(tk.END, "카카오뱅크 3333-03-1751818 (예금주: 김하늘)\n", ("donation_item", "donation_account"))
        self.program_info_text.insert(tk.END, "⚖️ 저작권 및 사용 조건\n", "section_title")
        self.program_info_text.insert(tk.END,
            "Copyright (c) 2023-2025 김하늘 (의문의돌맹이). All rights reserved.\n",
            "copyright_item"
        )
        self.program_info_text.insert(tk.END, "본 프로그램의 ", "copyright_item")
        self.program_info_text.insert(tk.END, "코드/구조 변경, 상업적 목적의 재배포, 무단 수정을 엄격히 금지", "copyright_notice")
        self.program_info_text.insert(tk.END, "합니다.\n", "copyright_item")
        self.program_info_text.insert(tk.END, "\n문의사항은 위 Discord로 연락주시면 감사하겠습니다.", "final_greeting")

    def _load_config(self):
        try:
            if os.path.exists(CONFIG_FILE_NAME):
                with open(CONFIG_FILE_NAME, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    loaded_path = config_data.get("excel_filepath")
                    if loaded_path and os.path.isabs(loaded_path):
                        self.current_excel_filepath = loaded_path
                    else:
                        print(f"'{CONFIG_FILE_NAME}'에 유효한 절대 경로가 없습니다. 기본 경로를 사용합니다.")
                        self._set_default_excel_path()
            else:
                print(f"'{CONFIG_FILE_NAME}'을 찾을 수 없습니다. 기본 경로를 사용합니다.")
                self._set_default_excel_path()
                self._save_config()
        except json.JSONDecodeError:
            print(f"'{CONFIG_FILE_NAME}' 파일 파싱 오류. 파일이 손상되었을 수 있습니다. 기본 경로를 사용합니다.")
            self._set_default_excel_path()
        except Exception as e:
            print(f"설정 파일 로드 중 예외 발생: {e}. 기본 경로를 사용합니다.")
            self._set_default_excel_path()
        
        self.excel_filepath_var.set(self.current_excel_filepath)

    def _set_default_excel_path(self):
        try:
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
        except NameError:
             base_path = os.getcwd()
        self.current_excel_filepath = os.path.join(base_path, DEFAULT_EXCEL_FILENAME)

    def _save_config(self):
        config_data = {"excel_filepath": self.current_excel_filepath}
        try:
            with open(CONFIG_FILE_NAME, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=4)
            print(f"설정이 '{CONFIG_FILE_NAME}'에 저장되었습니다: {self.current_excel_filepath}")
        except Exception as e:
            print(f"설정 파일 저장 중 오류 발생: {e}")
            messagebox.showerror("설정 저장 오류", f"설정 저장 중 오류 발생: {e}", parent=self.root if hasattr(self, 'root') else None)

    def _change_excel_path(self):
        global excel_manager
        
        initial_dir = os.path.dirname(self.current_excel_filepath) if self.current_excel_filepath else os.getcwd()
        initial_file = os.path.basename(self.current_excel_filepath) if self.current_excel_filepath else DEFAULT_EXCEL_FILENAME

        new_filepath = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=initial_file,
            title="엑셀 파일 저장 위치 선택",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )

        if new_filepath:
            new_abs_filepath = os.path.abspath(new_filepath)
            current_abs_filepath = os.path.abspath(self.current_excel_filepath) if self.current_excel_filepath else ""

            if new_abs_filepath != current_abs_filepath:
                self.current_excel_filepath = new_abs_filepath
                self.excel_filepath_var.set(self.current_excel_filepath)
                
                excel_manager = ExcelManager(self.current_excel_filepath)
                self.excel_manager = excel_manager
                
                self._save_config()
                self._update_settings_status_label(f"경로가 변경되었습니다: {self.current_excel_filepath}")
                
                if not os.path.exists(self.current_excel_filepath):
                    print(f"새 경로 '{self.current_excel_filepath}'에 파일이 없습니다. 템플릿 생성을 시도합니다.")
                    if self.excel_manager.create_excel_template():
                        messagebox.showinfo("템플릿 생성 완료", 
                                            f"새로운 경로에 '{os.path.basename(self.current_excel_filepath)}' 템플릿이 자동 생성되었습니다.",
                                            parent=self.root)
                        self._update_settings_status_label(f"새 경로에 템플릿 생성됨: {self.current_excel_filepath}")
                    else:
                        messagebox.showerror("템플릿 생성 실패", 
                                             f"새로운 경로에 템플릿 생성 실패: {self.current_excel_filepath}\n프로그램 로그를 확인해주세요.",
                                             parent=self.root)
                        self._update_settings_status_label(f"오류: 새 경로에 템플릿 생성 실패.", error=True)
                else:
                     messagebox.showinfo("경로 변경 완료", f"엑셀 파일 경로가 다음으로 설정되었습니다:\n{self.current_excel_filepath}", parent=self.root)
                
                # 사용 방법 탭의 파일명 즉시 업데이트
                self.usage_guide_text.config(state=tk.NORMAL)
                self._populate_usage_guide_tab_content() # 파일명 변경된 내용으로 업데이트
                self.usage_guide_text.config(state=tk.DISABLED)
            else:
                self._update_settings_status_label("경로가 변경되지 않았습니다 (기존 경로와 동일).")
        else:
            self._update_settings_status_label("경로 변경이 취소되었습니다.")

    def _update_settings_status_label(self, message, error=False):
        if hasattr(self, 'settings_status_label'):
            self.settings_status_label.config(text=message)
            if error:
                self.settings_status_label.config(style="Error.TLabel")
            else:
                self.settings_status_label.config(style="Placeholder.TLabel")


class HuntingResultForm:
    def __init__(self, parent_frame, main_app_ref):
        self.parent_frame = parent_frame
        self.main_app = main_app_ref
        self.canvas = tk.Canvas(parent_frame, highlightthickness=0, bg=DARK_BACKGROUND)
        self.scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.main_form_content_frame = ttk.Frame(self.scrollable_frame, padding="20")
        self.main_form_content_frame.pack(expand=True, fill=tk.BOTH)
        current_row_idx = 0

        # --- 입력 가이드라인 라벨 추가 ---
        hunt_guideline_text = (
            "💡 시작메소: 모든 준비(물약/지참비 등) 후 실제 보유액 | 판매후메소: 일반템 판매 후 최종액"
        )
        input_guideline_label_hunt = ttk.Label(
            self.main_form_content_frame,
            text=hunt_guideline_text,
            justify=tk.LEFT,
            style="Guide.TLabel",
            wraplength=700 
        )
        input_guideline_label_hunt.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=(0, 10), padx=5) # pady 상단 0, 하단 10
        Tooltip(input_guideline_label_hunt, "가장 중요한 입력 원칙입니다! 각 필드에 마우스를 올리면 더 자세한 설명을 볼 수 있어요.")
        current_row_idx += 1
        # --- 여기까지 가이드라인 라벨 추가 ---

        # --- 사냥 시작 정보 ---
        start_info_group = ttk.LabelFrame(self.main_form_content_frame, text="[ 사냥 시작 정보 ]", padding=(10, 5))
        start_info_group.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=(5,5), padx=5) # pady 조정
        start_info_group.columnconfigure(1, weight=1); start_info_group.columnconfigure(3, weight=1)
        s_current_row = 0
        
        lbl_date = ttk.Label(start_info_group, text="날짜:")
        lbl_date.grid(row=s_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.date_entry = ttk.Entry(start_info_group, width=20)
        self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.date_entry.grid(row=s_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        date_tooltip_text = "날짜 (YYYY-MM-DD). 'MMDD', 'YYMMDD' 등 입력 시 자동 완성."
        Tooltip(self.date_entry, date_tooltip_text); Tooltip(lbl_date, date_tooltip_text)
        self.date_entry.bind("<FocusOut>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))
        self.date_entry.bind("<Return>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))

        lbl_map_name = ttk.Label(start_info_group, text="맵명:")
        lbl_map_name.grid(row=s_current_row, column=2, padx=5, pady=5, sticky=tk.W)
        self.map_name_entry = ttk.Entry(start_info_group, width=20)
        self.map_name_entry.grid(row=s_current_row, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.map_name_entry, "사냥한 맵의 이름을 입력하세요 (필수)"); Tooltip(lbl_map_name, "사냥한 맵의 이름을 입력하세요 (필수)")
        s_current_row += 1

        lbl_start_time = ttk.Label(start_info_group, text="시작시간 (HH:MM):")
        lbl_start_time.grid(row=s_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_time_entry = ttk.Entry(start_info_group, width=20)
        self.start_time_entry.grid(row=s_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        time_tooltip_text = "시간 (HH:MM). 'HHMM' 또는 'HMM' 입력 시 자동 완성."
        Tooltip(self.start_time_entry, time_tooltip_text); Tooltip(lbl_start_time, time_tooltip_text)
        self.start_time_entry.bind("<FocusOut>", lambda event, entry=self.start_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        self.start_time_entry.bind("<Return>", lambda event, entry=self.start_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))

        lbl_start_meso = ttk.Label(start_info_group, text="시작 메소:")
        lbl_start_meso.grid(row=s_current_row, column=2, padx=5, pady=5, sticky=tk.W)
        self.start_meso_entry = ttk.Entry(start_info_group, width=20); self.start_meso_entry.insert(0, "0")
        self.start_meso_entry.grid(row=s_current_row, column=3, padx=5, pady=5, sticky=tk.EW)
        start_meso_tooltip = ("사냥 시작 직전, 캐릭터가 '실제로 보유 중인' 메소입니다.\n\n"
                              "※ 중요 ※\n"
                              "- 물약 구매, 맵 입장료(지참비) 등 모든 사전 준비 비용을 '이미 지출한 후'의 금액을 입력하세요.\n"
                              "- 예: 원래 100만 있었는데 물약 5만, 지참비 1만 썼다면 '94만' 입력.")
        Tooltip(self.start_meso_entry, start_meso_tooltip); Tooltip(lbl_start_meso, start_meso_tooltip)
        s_current_row += 1

        lbl_start_exp = ttk.Label(start_info_group, text="시작 경험치:")
        lbl_start_exp.grid(row=s_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_exp_entry = ttk.Entry(start_info_group, width=20); self.start_exp_entry.insert(0, "0")
        self.start_exp_entry.grid(row=s_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.start_exp_entry, "사냥 시작 시 경험치 수치"); Tooltip(lbl_start_exp, "사냥 시작 시 경험치 수치")

        lbl_entry_fee = ttk.Label(start_info_group, text="지참비:")
        lbl_entry_fee.grid(row=s_current_row, column=2, padx=5, pady=5, sticky=tk.W)
        self.entry_fee_entry = ttk.Entry(start_info_group, width=20); self.entry_fee_entry.insert(0, "0")
        self.entry_fee_entry.grid(row=s_current_row, column=3, padx=5, pady=5, sticky=tk.EW)
        entry_fee_tooltip = ("맵 입장료 등, 해당 사냥 세션을 위해 '고정적으로 지출된' 비용입니다.\n\n"
                             "※ 참고 ※\n"
                             "- 이 금액은 '시작 메소' 계산 시 이미 차감되었다고 가정합니다.\n"
                             "- 여기에 입력된 지참비는 '순수익' 계산 시 총수익에서 추가로 제외됩니다.")
        Tooltip(self.entry_fee_entry, entry_fee_tooltip); Tooltip(lbl_entry_fee, entry_fee_tooltip)
        current_row_idx += 1

        # --- 사냥 종료 정보 ---
        end_info_group = ttk.LabelFrame(self.main_form_content_frame, text="[ 사냥 종료 정보 ]", padding=(10, 5))
        end_info_group.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=10, padx=5)
        end_info_group.columnconfigure(1, weight=1); end_info_group.columnconfigure(3, weight=1)
        e_current_row = 0

        lbl_end_time = ttk.Label(end_info_group, text="종료시간 (HH:MM):")
        lbl_end_time.grid(row=e_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.end_time_entry = ttk.Entry(end_info_group, width=20)
        self.end_time_entry.grid(row=e_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.end_time_entry, time_tooltip_text); Tooltip(lbl_end_time, time_tooltip_text)
        self.end_time_entry.bind("<FocusOut>", lambda event, entry=self.end_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        self.end_time_entry.bind("<Return>", lambda event, entry=self.end_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        
        lbl_end_meso = ttk.Label(end_info_group, text="종료 메소:")
        lbl_end_meso.grid(row=e_current_row, column=2, padx=5, pady=5, sticky=tk.W)
        self.end_meso_entry = ttk.Entry(end_info_group, width=20); self.end_meso_entry.insert(0, "0")
        self.end_meso_entry.grid(row=e_current_row, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.end_meso_entry, "사냥 종료 시 보유 메소 (아이템 판매 전). '사냥메소수익' 계산에 사용됩니다."); Tooltip(lbl_end_meso, "사냥 종료 시 보유 메소 (아이템 판매 전). '사냥메소수익' 계산에 사용됩니다.")
        e_current_row += 1

        lbl_sold_meso = ttk.Label(end_info_group, text="판매 후 메소:")
        lbl_sold_meso.grid(row=e_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.sold_meso_entry = ttk.Entry(end_info_group, width=20); self.sold_meso_entry.insert(0, "0")
        self.sold_meso_entry.grid(row=e_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        sold_meso_tooltip = ("사냥 중 획득한 '일반 아이템(잡템)'을 모두 상점에 판매한 후, 최종적으로 캐릭터가 보유하게 된 메소입니다.\n\n"
                             "※ 주의 ※\n"
                             "- '고가 아이템' 판매액은 포함하지 마세요 (별도 목록으로 관리).\n"
                             "- '종료 메소'에서 일반템 판매로 늘어난 금액을 정확히 반영해주세요.\n"
                             "- 미입력 시 '종료 메소'와 동일하게 처리 (일반템 수익 0).")
        Tooltip(self.sold_meso_entry, sold_meso_tooltip); Tooltip(lbl_sold_meso, sold_meso_tooltip)
        
        lbl_end_exp = ttk.Label(end_info_group, text="종료 경험치:")
        lbl_end_exp.grid(row=e_current_row, column=2, padx=5, pady=5, sticky=tk.W)
        self.end_exp_entry = ttk.Entry(end_info_group, width=20); self.end_exp_entry.insert(0, "0")
        self.end_exp_entry.grid(row=e_current_row, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.end_exp_entry, "사냥 종료 시 경험치 수치"); Tooltip(lbl_end_exp, "사냥 종료 시 경험치 수치")
        e_current_row += 1

        lbl_coupon = ttk.Label(end_info_group, text="15분 쿠폰 수:")
        lbl_coupon.grid(row=e_current_row, column=0, padx=5, pady=5, sticky=tk.W)
        self.coupon_entry = ttk.Entry(end_info_group, width=20); self.coupon_entry.insert(0, "0")
        self.coupon_entry.grid(row=e_current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.coupon_entry, "사용한 15분 경험치 2배 쿠폰의 개수 (원경험치 계산용)"); Tooltip(lbl_coupon, "사용한 15분 경험치 2배 쿠폰의 개수 (원경험치 계산용)")
        current_row_idx += 1

        # --- 고가 아이템 목록 ---
        self.rare_items_outer_frame = ttk.LabelFrame(self.main_form_content_frame, text="획득한 고가 아이템 목록", padding=10)
        self.rare_items_outer_frame.grid(row=current_row_idx, column=0, columnspan=4, sticky='ew', pady=(10,0), padx=5)
        
        self.rare_items_manager = DynamicEntryListFrame(
            parent_container=self.rare_items_outer_frame,
            item_type_name_kor="고가 아이템",
            fields_config_list=RARE_ITEM_FIELDS_CONFIG, # 툴팁은 CONSUMABLE_ITEM_FIELDS_CONFIG 에서 이미 수정됨
            no_item_text_format_str="등록된 {}이(가) 없습니다.",
            canvas_for_scroll_update=self.canvas,
            example_placeholder_str="예: 혼돈의 파편",
            item_label_frame_parent=self.parent_frame
        )
        self.rare_items_manager.pack(fill=tk.X, expand=True, pady=(0, 5))
        current_row_idx +=1

        add_rare_item_btn_frame = ttk.Frame(self.main_form_content_frame)
        add_rare_item_btn_frame.grid(row=current_row_idx, column=0, columnspan=4, pady=(0,5), sticky=tk.W, padx=15)
        self.add_rare_item_button = ttk.Button(add_rare_item_btn_frame, text="고가 아이템 추가 (+)", command=self.rare_items_manager.add_new_item_entry_row)
        self.add_rare_item_button.pack(side=tk.LEFT)
        Tooltip(self.add_rare_item_button, "획득한 고가 아이템 정보를 추가합니다.")
        current_row_idx += 1
        
        # --- 소모/획득 아이템 목록 ---
        self.consumables_outer_frame = ttk.LabelFrame(self.main_form_content_frame, text="소모/획득 아이템 목록", padding=10)
        self.consumables_outer_frame.grid(row=current_row_idx, column=0, columnspan=4, sticky='ew', pady=(10,0), padx=5)

        self.consumables_manager = DynamicEntryListFrame(
            parent_container=self.consumables_outer_frame,
            item_type_name_kor="소모/획득 아이템",
            fields_config_list=CONSUMABLE_ITEM_FIELDS_CONFIG, # 툴팁 강화된 설정 사용
            no_item_text_format_str="등록된 {}이(가) 없습니다.",
            canvas_for_scroll_update=self.canvas,
            example_placeholder_str="예: 파워엘릭서",
            item_label_frame_parent=self.parent_frame
        )
        self.consumables_manager.pack(fill=tk.X, expand=True, pady=(0,5))
        current_row_idx += 1

        add_consumable_btn_frame = ttk.Frame(self.main_form_content_frame)
        add_consumable_btn_frame.grid(row=current_row_idx, column=0, columnspan=4, pady=(0,5), sticky=tk.W, padx=15)
        self.add_consumable_button = ttk.Button(add_consumable_btn_frame, text="소모/획득 아이템 추가 (+)", command=self.consumables_manager.add_new_item_entry_row)
        self.add_consumable_button.pack(side=tk.LEFT)
        Tooltip(self.add_consumable_button, "사용한 소모 아이템 또는 획득한 소모 아이템 정보를 추가합니다.")
        current_row_idx += 1

        # --- 버튼 프레임 ---
        self.action_buttons_frame = ttk.Frame(self.main_form_content_frame)
        self.action_buttons_frame.grid(row=current_row_idx, column=0, columnspan=4, pady=20)
        self.save_button = ttk.Button(self.action_buttons_frame, text="저장 및 계산", command=self.save_data)
        self.save_button.pack(side=tk.LEFT, padx=10)
        Tooltip(self.save_button, "입력한 사냥 결과를 저장하고 계산합니다.")
        self.clear_button = ttk.Button(self.action_buttons_frame, text="초기화", command=self.clear_fields)
        self.clear_button.pack(side=tk.LEFT, padx=10)
        Tooltip(self.clear_button, "모든 입력 필드를 초기화합니다.")

        self.main_form_content_frame.columnconfigure(0, weight=0)
        self.main_form_content_frame.columnconfigure(1, weight=1)
        self.main_form_content_frame.columnconfigure(2, weight=0)
        self.main_form_content_frame.columnconfigure(3, weight=1)

    def _on_canvas_configure(self, event):
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)
        self.scrollable_frame.configure(width=canvas_width)
    
    def _get_all_rare_items_data(self):
        items_list_for_excel = []
        total_rare_item_value = 0
        all_rare_item_widgets = self.rare_items_manager.get_all_entry_widgets()
        msg_parent = self.rare_items_manager.get_messagebox_parent()

        for item_widgets_set in all_rare_item_widgets:
            try:
                name_widget = item_widgets_set['name_entry']
                value_widget = item_widgets_set['value_entry']
                
                item_name = MainApplication.static_get_entry_value(
                    name_widget, str, "", field_name_for_error="고가템명", parent_for_msgbox=msg_parent
                )
                if not item_name: continue

                item_value = MainApplication.static_get_entry_value(
                    value_widget, int, "0", field_name_for_error=f"'{item_name}' 가치", parent_for_msgbox=msg_parent
                )

                if item_name and item_value > 0:
                    items_list_for_excel.append((item_name, item_value))
                    total_rare_item_value += item_value
                elif item_name and item_value <= 0:
                    messagebox.showwarning("고가템 가치 오류", f"고가 아이템 '{item_name}'의 가치가 0 이하입니다. 계산에서 제외됩니다.", parent=msg_parent, detail="아이템 가치는 0보다 커야 합니다.")
            except ValueError:
                continue
        return items_list_for_excel, total_rare_item_value

    def _process_consumable_items(self):
        total_consumable_cost = 0
        total_consumable_gained_profit = 0
        all_consumable_widgets = self.consumables_manager.get_all_entry_widgets()
        msg_parent = self.consumables_manager.get_messagebox_parent()

        for item_widgets_set in all_consumable_widgets:
            try:
                name_widget = item_widgets_set['name_entry']
                price_value_widget = item_widgets_set['price_value_entry']
                start_qty_widget = item_widgets_set['start_qty_entry']
                end_qty_widget = item_widgets_set['end_qty_entry']

                item_name = MainApplication.static_get_entry_value(
                    name_widget, str, "", field_name_for_error="소모/획득템 명칭", parent_for_msgbox=msg_parent
                )
                if not item_name: continue

                item_price_or_value = MainApplication.static_get_entry_value(
                    price_value_widget, int, "0", field_name_for_error=f"'{item_name}' 개당 가격/가치", parent_for_msgbox=msg_parent
                )
                start_qty = MainApplication.static_get_entry_value(
                    start_qty_widget, int, "0", field_name_for_error=f"'{item_name}' 시작 개수", parent_for_msgbox=msg_parent
                )
                end_qty = MainApplication.static_get_entry_value(
                    end_qty_widget, int, "0", field_name_for_error=f"'{item_name}' 종료 개수", parent_for_msgbox=msg_parent
                )

                quantity_change = end_qty - start_qty

                if quantity_change < 0: 
                    used_quantity = -quantity_change
                    item_cost = used_quantity * item_price_or_value
                    total_consumable_cost += item_cost
                elif quantity_change > 0: 
                    gained_quantity = quantity_change
                    item_gained_profit = gained_quantity * item_price_or_value
                    total_consumable_gained_profit += item_gained_profit
            except ValueError:
                continue
        return total_consumable_cost, total_consumable_gained_profit

    def calculate_hunting_data(self, data_dict):
        calculated_data = data_dict.copy()
        total_hunting_minutes = 0
        try:
            start_dt_str = data_dict.get('시작시간'); end_dt_str = data_dict.get('종료시간')
            if not start_dt_str : messagebox.showerror("계산 오류", "사냥 시작시간이 입력되지 않았습니다.", parent=self.parent_frame); raise ValueError("사냥 시작시간 누락")
            if not end_dt_str : messagebox.showerror("계산 오류", "사냥 종료시간이 입력되지 않았습니다.", parent=self.parent_frame); raise ValueError("사냥 종료시간 누락")
            
            start_dt = datetime.datetime.strptime(start_dt_str, "%H:%M")
            end_dt = datetime.datetime.strptime(end_dt_str, "%H:%M")

            if end_dt < start_dt: end_dt += datetime.timedelta(days=1)
            duration_delta = end_dt - start_dt
            total_hunting_minutes = int(duration_delta.total_seconds() / 60)

            if total_hunting_minutes < 0 : 
                messagebox.showerror("시간 오류", "사냥 시간이 음수입니다. 시작/종료 시간을 확인해주세요.", parent=self.parent_frame)
                raise ValueError("사냥 시간 < 0")
        except ValueError as e:
            if "누락" not in str(e) and "시간 < 0" not in str(e) and "잘못된 시간 형식" not in str(e) and "HH:MM" not in str(e): 
                 messagebox.showerror("시간 형식 오류", "시작시간 또는 종료시간 형식이 잘못되었습니다 (HH:MM).", parent=self.parent_frame)
            raise 

        calculated_data['사냥메소수익'] = data_dict.get('종료메소',0) - data_dict.get('시작메소',0)
        
        sold_meso = data_dict.get('판매후메소', data_dict.get('종료메소',0))
        end_meso = data_dict.get('종료메소',0)
        if sold_meso >= end_meso:
            calculated_data['일반템수익'] = sold_meso - end_meso
        else:
            messagebox.showwarning("메소 입력 오류", "판매 후 메소가 종료 메소보다 적습니다. 일반템 수익은 0으로 처리됩니다.", parent=self.parent_frame)
            calculated_data['일반템수익'] = 0
            
        calculated_data['경험치수익'] = data_dict.get('종료경험치',0) - data_dict.get('시작경험치',0)
        
        total_gained_exp = calculated_data['경험치수익']
        coupon_count = data_dict.get('15분쿠폰사용횟수', 0)
        calculated_data['원경험치'] = 0

        if total_hunting_minutes > 0:
            if total_gained_exp > 0:
                actual_coupon_minutes = min(coupon_count * 15, total_hunting_minutes)
                non_coupon_minutes = total_hunting_minutes - actual_coupon_minutes
                coupon_exp_rate = 2.0
                effective_denominator_minutes = non_coupon_minutes + (actual_coupon_minutes * coupon_exp_rate)
                if effective_denominator_minutes > 0:
                    base_exp_per_effective_minute = total_gained_exp / effective_denominator_minutes
                    calculated_data['원경험치'] = round(base_exp_per_effective_minute * total_hunting_minutes)
                else:
                    if coupon_count > 0 and total_hunting_minutes <= coupon_count * 15 and coupon_exp_rate > 0:
                        calculated_data['원경험치'] = round(total_gained_exp / coupon_exp_rate)
                    else:
                        calculated_data['원경험치'] = total_gained_exp 
        elif total_gained_exp > 0 and total_hunting_minutes <= 0 :
             messagebox.showwarning("경험치 계산 경고", "사냥 시간이 0분 이하이지만 경험치를 얻었습니다. 원 경험치는 0으로 처리됩니다.", parent=self.parent_frame)
        
        rare_items_list, total_rare_value = self._get_all_rare_items_data()
        calculated_data['rare_items_list'] = rare_items_list
        calculated_data['고가아이템가치'] = total_rare_value
        calculated_data['고가아이템'] = ", ".join([item[0] for item in rare_items_list]) if rare_items_list else ""

        total_consumable_cost, total_consumable_gained_profit = self._process_consumable_items()
        calculated_data['소모아이템비'] = total_consumable_cost
        calculated_data['소모템획득수익'] = total_consumable_gained_profit
        
        calculated_data['총수익'] = (calculated_data.get('사냥메소수익',0) + 
                                   calculated_data.get('일반템수익',0) + 
                                   total_rare_value +
                                   total_consumable_gained_profit)
        calculated_data['순수익'] = calculated_data.get('총수익',0) - data_dict.get('지참비',0) - calculated_data.get('소모아이템비',0)
        
        return calculated_data

    def save_data(self):
        raw_data_from_form = {}
        try:
            excel_file_path = self.main_app.excel_manager.filename
            if not os.path.exists(excel_file_path):
                print(f"'{os.path.basename(excel_file_path)}' 파일이 없어 템플릿 생성을 시도합니다. (HuntingResultForm.save_data)")
                if not self.main_app.excel_manager.create_excel_template():
                    messagebox.showerror("템플릿 생성 실패", 
                                         f"'{os.path.basename(excel_file_path)}' 템플릿 생성에 실패하여 저장할 수 없습니다.\n"
                                         "프로그램을 재시작하거나 파일 경로 설정을 확인해주세요.", 
                                         parent=self.parent_frame)
                    return
                else:
                    messagebox.showinfo("템플릿 자동 생성", 
                                        f"'{os.path.basename(excel_file_path)}' 파일이 없어 새로 생성했습니다.\n이제 데이터를 저장합니다.", 
                                        parent=self.parent_frame)
                    self.parent_frame.update_idletasks() 
                    self.parent_frame.after(100, lambda: self.map_name_entry.focus_set()) 
            
            raw_data_from_form['날짜'] = MainApplication.static_get_entry_value(self.date_entry, str, field_name_for_error="날짜", parent_for_msgbox=self.parent_frame)
            if not raw_data_from_form['날짜']: messagebox.showerror("입력 오류", "날짜를 입력해주세요.", parent=self.parent_frame); self.date_entry.focus_set(); return
            
            raw_data_from_form['맵명'] = MainApplication.static_get_entry_value(self.map_name_entry, str, field_name_for_error="맵명", parent_for_msgbox=self.parent_frame)
            if not raw_data_from_form['맵명']: messagebox.showerror("입력 오류", "맵명을 입력해주세요.", parent=self.parent_frame); self.map_name_entry.focus_set(); return
            
            raw_data_from_form['시작시간'] = self.start_time_entry.get().strip()
            raw_data_from_form['종료시간'] = self.end_time_entry.get().strip()

            if not raw_data_from_form['시작시간']: messagebox.showerror("입력 오류", "시작시간을 입력해주세요.", parent=self.parent_frame); self.start_time_entry.focus_set(); return
            if not raw_data_from_form['종료시간']: messagebox.showerror("입력 오류", "종료시간을 입력해주세요.", parent=self.parent_frame); self.end_time_entry.focus_set(); return
            
            try: 
                datetime.datetime.strptime(raw_data_from_form['시작시간'], "%H:%M")
                datetime.datetime.strptime(raw_data_from_form['종료시간'], "%H:%M")
            except ValueError:
                messagebox.showerror("입력 오류", "시작시간 또는 종료시간 형식이 올바르지 않습니다 (HH:MM).", parent=self.parent_frame); return

            raw_data_from_form['시작메소'] = MainApplication.static_get_entry_value(self.start_meso_entry, int, "0", field_name_for_error="시작 메소", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['종료메소'] = MainApplication.static_get_entry_value(self.end_meso_entry, int, "0", field_name_for_error="종료 메소", parent_for_msgbox=self.parent_frame)
            
            sold_meso_str = self.sold_meso_entry.get().strip()
            if not sold_meso_str or sold_meso_str == "0":
                raw_data_from_form['판매후메소'] = raw_data_from_form['종료메소']
            else:
                raw_data_from_form['판매후메소'] = MainApplication.static_get_entry_value(
                    self.sold_meso_entry, int, str(raw_data_from_form['종료메소']), 
                    field_name_for_error="판매 후 메소", parent_for_msgbox=self.parent_frame
                )
            
            raw_data_from_form['시작경험치'] = MainApplication.static_get_entry_value(self.start_exp_entry, int, "0", field_name_for_error="시작 경험치", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['종료경험치'] = MainApplication.static_get_entry_value(self.end_exp_entry, int, "0", field_name_for_error="종료 경험치", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['15분쿠폰사용횟수'] = MainApplication.static_get_entry_value(self.coupon_entry, int, "0", field_name_for_error="15분 쿠폰 수", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['지참비'] = MainApplication.static_get_entry_value(self.entry_fee_entry, int, "0", field_name_for_error="지참비", parent_for_msgbox=self.parent_frame)
            
            final_data_to_save = self.calculate_hunting_data(raw_data_from_form)
            
            if self.main_app.excel_manager.save_data_to_sheet(final_data_to_save, "사냥세션", HUNTING_SHEET_COLUMNS, "사냥", parent_for_msgbox=self.parent_frame):
                date_to_update = final_data_to_save.get('날짜')
                hunting_meso_val = final_data_to_save.get('사냥메소수익', 0)
                rare_item_val = final_data_to_save.get('고가아이템가치', 0)
                normal_item_val = final_data_to_save.get('일반템수익', 0)
                consumable_gained_val = final_data_to_save.get('소모템획득수익', 0)
                consumable_cost_val = final_data_to_save.get('소모아이템비', 0)
                entry_fee_val = final_data_to_save.get('지참비', 0)

                self.main_app.excel_manager.update_daily_summary(
                    date_str=date_to_update,
                    hunting_meso=hunting_meso_val,
                    rare_item_profit=rare_item_val,
                    normal_item_profit=normal_item_val,
                    consumable_gained_profit=consumable_gained_val,
                    consumable_cost=consumable_cost_val,
                    entry_fee=entry_fee_val,
                    parent_for_msgbox=self.parent_frame
                )
                map_name_to_update = final_data_to_save.get('맵명')
                session_main_profit_hunt = final_data_to_save.get('사냥메소수익', 0) + final_data_to_save.get('일반템수익', 0)
                session_rare_profit_hunt = final_data_to_save.get('고가아이템가치', 0)
                session_consumable_gained_hunt = final_data_to_save.get('소모템획득수익', 0)

                self.main_app.excel_manager.update_map_summary(
                    map_name=map_name_to_update,
                    session_type="사냥",
                    profit_from_session=session_main_profit_hunt,
                    rare_item_profit_from_session=session_rare_profit_hunt,
                    consumable_gained_profit_from_session=session_consumable_gained_hunt,
                    parent_for_msgbox=self.parent_frame
                )
                
                date_for_weekday_update = final_data_to_save.get('날짜')
                session_type_hunt = "사냥"
                val_pure_revenue_hunt = final_data_to_save.get('사냥메소수익', 0)
                val_rare_item_hunt = final_data_to_save.get('고가아이템가치', 0)
                val_normal_item_hunt = final_data_to_save.get('일반템수익', 0)
                val_consumable_gained_hunt = final_data_to_save.get('소모템획득수익', 0)
                val_net_profit_hunt = final_data_to_save.get('순수익', 0) 

                self.main_app.excel_manager.update_weekday_summary(
                    date_str=date_for_weekday_update,
                    session_type=session_type_hunt,
                    val_pure_revenue_contribution=val_pure_revenue_hunt,
                    val_rare_item_contribution=val_rare_item_hunt,
                    val_normal_item_contribution=val_normal_item_hunt,
                    val_consumable_gained_contribution=val_consumable_gained_hunt,
                    val_net_profit_contribution=val_net_profit_hunt,
                    parent_for_msgbox=self.parent_frame
                )
                current_filename = os.path.basename(self.main_app.excel_manager.filename)
                messagebox.showinfo("저장 완료", f"사냥 결과가 '{current_filename}'에 저장되고 관련 통계가 업데이트되었습니다.", parent=self.parent_frame)
                self.clear_fields()
        
        except ValueError as e: 
            print(f"데이터 입력/계산 오류로 저장 중단 (HuntingResultForm): {e}")
            return 
        except Exception as e:
            messagebox.showerror("알 수 없는 오류", f"데이터 저장 중 예기치 않은 오류 발생: {e}", parent=self.parent_frame)
            print(f"알 수 없는 오류 (HuntingResultForm.save_data): {e}")
            import traceback; traceback.print_exc()
            return

    def clear_fields(self):
        self.date_entry.delete(0, tk.END); self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.map_name_entry.delete(0, tk.END)
        self.start_time_entry.delete(0, tk.END); self.end_time_entry.delete(0, tk.END)
        self.start_meso_entry.delete(0, tk.END); self.start_meso_entry.insert(0, "0")
        self.end_meso_entry.delete(0, tk.END); self.end_meso_entry.insert(0, "0")
        self.sold_meso_entry.delete(0, tk.END); self.sold_meso_entry.insert(0, "0")
        self.start_exp_entry.delete(0, tk.END); self.start_exp_entry.insert(0, "0")
        self.end_exp_entry.delete(0, tk.END); self.end_exp_entry.insert(0, "0")
        self.coupon_entry.delete(0, tk.END); self.coupon_entry.insert(0, "0")
        self.entry_fee_entry.delete(0, tk.END); self.entry_fee_entry.insert(0, "0")
        
        self.rare_items_manager.clear_all_item_entries()
        self.consumables_manager.clear_all_item_entries()
        
        self.map_name_entry.focus_set()
        print("사냥 결과 필드 초기화됨.")
        
        self.canvas.after_idle(lambda: (
            self.canvas.configure(scrollregion=self.canvas.bbox("all")),
            self.canvas.yview_moveto(0)
        ))


class JjulResultForm:
    def __init__(self, parent_frame, main_app_ref):
        self.parent_frame = parent_frame
        self.main_app = main_app_ref
        self.canvas = tk.Canvas(parent_frame, highlightthickness=0, bg=DARK_BACKGROUND)
        self.scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set); self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y"); self.canvas.bind("<Configure>", self._on_canvas_configure)
        
        self.main_form_content_frame = ttk.Frame(self.scrollable_frame, padding="20"); 
        self.main_form_content_frame.pack(expand=True, fill=tk.BOTH)
        current_row_idx = 0

        # --- 입력 가이드라인 라벨 추가 ---
        jjul_guideline_text = (
            "💡 시작메소: 모든 준비(물약 등) 후 실제 보유액 | 판매후메소: 일반템 판매 후 최종액"
        )
        input_guideline_label_jjul = ttk.Label(
            self.main_form_content_frame,
            text=jjul_guideline_text,
            justify=tk.LEFT,
            style="Guide.TLabel",
            wraplength=700
        )
        input_guideline_label_jjul.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=(0, 10), padx=5)
        Tooltip(input_guideline_label_jjul, "가장 중요한 입력 원칙입니다! 각 필드에 마우스를 올리면 더 자세한 설명을 볼 수 있어요.")
        current_row_idx += 1
        # --- 여기까지 가이드라인 라벨 추가 ---

        # --- 쩔 시작 정보 ---
        start_info_group_jjul = ttk.LabelFrame(self.main_form_content_frame, text="[ 쩔 시작 정보 ]", padding=(10, 5))
        start_info_group_jjul.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=(5,5), padx=5) # pady 조정
        start_info_group_jjul.columnconfigure(1, weight=1); start_info_group_jjul.columnconfigure(3, weight=1)
        s_current_row_j = 0

        lbl_j_date = ttk.Label(start_info_group_jjul, text="날짜:")
        lbl_j_date.grid(row=s_current_row_j, column=0, padx=5, pady=5, sticky=tk.W)
        self.date_entry = ttk.Entry(start_info_group_jjul, width=20)
        self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.date_entry.grid(row=s_current_row_j, column=1, padx=5, pady=5, sticky=tk.EW)
        date_tooltip_text = "날짜 (YYYY-MM-DD). 'MMDD', 'YYMMDD' 등 입력 시 자동 완성."
        Tooltip(self.date_entry, date_tooltip_text); Tooltip(lbl_j_date, date_tooltip_text)
        self.date_entry.bind("<FocusOut>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))
        self.date_entry.bind("<Return>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))

        lbl_j_map = ttk.Label(start_info_group_jjul, text="맵명:")
        lbl_j_map.grid(row=s_current_row_j, column=2, padx=5, pady=5, sticky=tk.W)
        self.map_name_entry = ttk.Entry(start_info_group_jjul, width=20)
        self.map_name_entry.grid(row=s_current_row_j, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.map_name_entry, "쩔을 진행한 맵의 이름 (필수)"); Tooltip(lbl_j_map, "쩔을 진행한 맵의 이름 (필수)")
        s_current_row_j += 1

        lbl_j_stime = ttk.Label(start_info_group_jjul, text="시작시간 (HH:MM):")
        lbl_j_stime.grid(row=s_current_row_j, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_time_entry = ttk.Entry(start_info_group_jjul, width=20)
        self.start_time_entry.grid(row=s_current_row_j, column=1, padx=5, pady=5, sticky=tk.EW)
        time_tooltip_text = "시간 (HH:MM). 'HHMM' 또는 'HMM' 입력 시 자동 완성." 
        Tooltip(self.start_time_entry, time_tooltip_text); Tooltip(lbl_j_stime, time_tooltip_text)
        self.start_time_entry.bind("<FocusOut>", lambda event, entry=self.start_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        self.start_time_entry.bind("<Return>", lambda event, entry=self.start_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        
        lbl_j_smeso = ttk.Label(start_info_group_jjul, text="시작 메소:")
        lbl_j_smeso.grid(row=s_current_row_j, column=2, padx=5, pady=5, sticky=tk.W)
        self.start_meso_entry = ttk.Entry(start_info_group_jjul, width=20); self.start_meso_entry.insert(0, "0")
        self.start_meso_entry.grid(row=s_current_row_j, column=3, padx=5, pady=5, sticky=tk.EW)
        # 툴팁은 HuntingResultForm과 동일한 내용 사용 (이미 강화됨)
        start_meso_tooltip_jjul = ("쩔 시작 직전, 캐릭터가 '실제로 보유 중인' 메소입니다.\n\n"
                                   "※ 중요 ※\n"
                                   "- 물약 구매 등 모든 사전 준비 비용을 '이미 지출한 후'의 금액을 입력하세요.\n"
                                   "- 예: 원래 100만 있었는데 물약 5만 썼다면 '95만' 입력.")
        Tooltip(self.start_meso_entry, start_meso_tooltip_jjul); Tooltip(lbl_j_smeso, start_meso_tooltip_jjul)
        s_current_row_j += 1

        lbl_j_party = ttk.Label(start_info_group_jjul, text="쩔 인원 수:")
        lbl_j_party.grid(row=s_current_row_j, column=0, padx=5, pady=5, sticky=tk.W)
        self.party_size_entry = ttk.Entry(start_info_group_jjul, width=20); self.party_size_entry.insert(0, "0")
        self.party_size_entry.grid(row=s_current_row_j, column=1, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.party_size_entry, "쩔을 받은 총 인원 수 (본인 제외)"); Tooltip(lbl_j_party, "쩔을 받은 총 인원 수 (본인 제외)")

        lbl_j_price = ttk.Label(start_info_group_jjul, text="1인당 쩔비:")
        lbl_j_price.grid(row=s_current_row_j, column=2, padx=5, pady=5, sticky=tk.W)
        self.price_per_person_entry = ttk.Entry(start_info_group_jjul, width=20); self.price_per_person_entry.insert(0, "0")
        self.price_per_person_entry.grid(row=s_current_row_j, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.price_per_person_entry, "쩔 손님 1명에게 받은 메소 (수수료 제외 전 금액 가능)"); Tooltip(lbl_j_price, "쩔 손님 1명에게 받은 메소 (수수료 제외 전 금액 가능)")
        current_row_idx += 1

        # --- 쩔 종료 정보 ---
        end_info_group_jjul = ttk.LabelFrame(self.main_form_content_frame, text="[ 쩔 종료 정보 ]", padding=(10, 5))
        end_info_group_jjul.grid(row=current_row_idx, column=0, columnspan=4, sticky="ew", pady=10, padx=5)
        end_info_group_jjul.columnconfigure(1, weight=1); end_info_group_jjul.columnconfigure(3, weight=1)
        e_current_row_j = 0
        
        lbl_j_etime = ttk.Label(end_info_group_jjul, text="종료시간 (HH:MM):")
        lbl_j_etime.grid(row=e_current_row_j, column=0, padx=5, pady=5, sticky=tk.W)
        self.end_time_entry = ttk.Entry(end_info_group_jjul, width=20)
        self.end_time_entry.grid(row=e_current_row_j, column=1, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.end_time_entry, time_tooltip_text); Tooltip(lbl_j_etime, time_tooltip_text)
        self.end_time_entry.bind("<FocusOut>", lambda event, entry=self.end_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        self.end_time_entry.bind("<Return>", lambda event, entry=self.end_time_entry: MainApplication.static_format_time_entry(event, entry, self.parent_frame))
        
        lbl_j_emeso = ttk.Label(end_info_group_jjul, text="종료 메소:")
        lbl_j_emeso.grid(row=e_current_row_j, column=2, padx=5, pady=5, sticky=tk.W)
        self.end_meso_entry = ttk.Entry(end_info_group_jjul, width=20); self.end_meso_entry.insert(0, "0")
        self.end_meso_entry.grid(row=e_current_row_j, column=3, padx=5, pady=5, sticky=tk.EW)
        Tooltip(self.end_meso_entry, "쩔 종료 시 보유 메소 (아이템 판매 전, 쩔비 받은 후). '일반템수익' 계산 시 중요."); Tooltip(lbl_j_emeso, "쩔 종료 시 보유 메소 (아이템 판매 전, 쩔비 받은 후). '일반템수익' 계산 시 중요.")
        e_current_row_j += 1

        lbl_j_smeso_after = ttk.Label(end_info_group_jjul, text="판매 후 메소:")
        lbl_j_smeso_after.grid(row=e_current_row_j, column=0, padx=5, pady=5, sticky=tk.W)
        self.sold_meso_entry = ttk.Entry(end_info_group_jjul, width=20); self.sold_meso_entry.insert(0, "0")
        self.sold_meso_entry.grid(row=e_current_row_j, column=1, padx=5, pady=5, sticky=tk.EW)
        # 툴팁은 HuntingResultForm과 동일한 내용 사용 (이미 강화됨)
        sold_meso_tooltip_jjul = ("쩔 중 획득한 '일반 아이템(잡템)'을 모두 상점에 판매한 후, 최종적으로 캐릭터가 보유하게 된 메소입니다.\n\n"
                                  "※ 주의 ※\n"
                                  "- '고가 아이템' 판매액은 포함하지 마세요 (별도 목록으로 관리).\n"
                                  "- '종료 메소'에서 일반템 판매로 늘어난 금액을 정확히 반영해주세요.\n"
                                  "- 미입력 시 '종료 메소'와 동일하게 처리 (일반템 수익 0).")
        Tooltip(self.sold_meso_entry, sold_meso_tooltip_jjul); Tooltip(lbl_j_smeso_after, sold_meso_tooltip_jjul)
        current_row_idx +=1

        # --- 고가 아이템 목록 (쩔) ---
        self.rare_items_outer_frame_jjul = ttk.LabelFrame(self.main_form_content_frame, text="획득한 고가 아이템 목록", padding=10)
        self.rare_items_outer_frame_jjul.grid(row=current_row_idx, column=0, columnspan=4, sticky='ew', pady=(10,0), padx=5)
        
        self.rare_items_manager_jjul = DynamicEntryListFrame(
            parent_container=self.rare_items_outer_frame_jjul,
            item_type_name_kor="고가 아이템",
            fields_config_list=RARE_ITEM_FIELDS_CONFIG,
            no_item_text_format_str="등록된 {}이(가) 없습니다.",
            canvas_for_scroll_update=self.canvas,
            example_placeholder_str="예: 뒤틀린 낙인의 영혼석",
            item_label_frame_parent=self.parent_frame
        )
        self.rare_items_manager_jjul.pack(fill=tk.X, expand=True, pady=(0,5))
        current_row_idx += 1
        
        add_rare_item_btn_frame_jjul = ttk.Frame(self.main_form_content_frame)
        add_rare_item_btn_frame_jjul.grid(row=current_row_idx, column=0, columnspan=4, pady=(0,5), sticky=tk.W, padx=15)
        self.add_rare_item_button_jjul = ttk.Button(add_rare_item_btn_frame_jjul, text="고가 아이템 추가 (+)", command=self.rare_items_manager_jjul.add_new_item_entry_row)
        self.add_rare_item_button_jjul.pack(side=tk.LEFT)
        Tooltip(self.add_rare_item_button_jjul, "쩔 중 획득한 고가 아이템 정보를 추가합니다.")
        current_row_idx += 1

        # --- 소모/획득 아이템 목록 (쩔) ---
        self.consumables_outer_frame_jjul = ttk.LabelFrame(self.main_form_content_frame, text="소모/획득 아이템 목록", padding=10)
        self.consumables_outer_frame_jjul.grid(row=current_row_idx, column=0, columnspan=4, sticky='ew', pady=(10,0), padx=5)

        self.consumables_manager_jjul = DynamicEntryListFrame(
            parent_container=self.consumables_outer_frame_jjul,
            item_type_name_kor="소모/획득 아이템",
            fields_config_list=CONSUMABLE_ITEM_FIELDS_CONFIG, # 툴팁 강화된 설정 사용
            no_item_text_format_str="등록된 {}이(가) 없습니다.",
            canvas_for_scroll_update=self.canvas,
            example_placeholder_str="예: 만병통치약",
            item_label_frame_parent=self.parent_frame
        )
        self.consumables_manager_jjul.pack(fill=tk.X, expand=True, pady=(0,5))
        current_row_idx += 1

        add_consumable_btn_frame_jjul = ttk.Frame(self.main_form_content_frame)
        add_consumable_btn_frame_jjul.grid(row=current_row_idx, column=0, columnspan=4, pady=(0,5), sticky=tk.W, padx=15)
        self.add_consumable_button_jjul = ttk.Button(add_consumable_btn_frame_jjul, text="소모/획득 아이템 추가 (+)", command=self.consumables_manager_jjul.add_new_item_entry_row)
        self.add_consumable_button_jjul.pack(side=tk.LEFT)
        Tooltip(self.add_consumable_button_jjul, "쩔 중 사용하거나 획득한 소모 아이템 정보를 추가합니다.")
        current_row_idx += 1

        # --- 버튼 프레임 (쩔) ---
        self.action_buttons_frame_jjul = ttk.Frame(self.main_form_content_frame)
        self.action_buttons_frame_jjul.grid(row=current_row_idx, column=0, columnspan=4, pady=20)
        self.save_button_jjul = ttk.Button(self.action_buttons_frame_jjul, text="저장 및 계산", command=self.save_data)
        self.save_button_jjul.pack(side=tk.LEFT, padx=10); Tooltip(self.save_button_jjul, "입력한 쩔 결과를 저장하고 계산합니다.")
        self.clear_button_jjul = ttk.Button(self.action_buttons_frame_jjul, text="초기화", command=self.clear_fields)
        self.clear_button_jjul.pack(side=tk.LEFT, padx=10); Tooltip(self.clear_button_jjul, "모든 입력 필드를 초기화합니다.")
        
        self.main_form_content_frame.columnconfigure(0, weight=0); self.main_form_content_frame.columnconfigure(1, weight=1)
        self.main_form_content_frame.columnconfigure(2, weight=0); self.main_form_content_frame.columnconfigure(3, weight=1)


    def _on_canvas_configure(self, event):
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)
        self.scrollable_frame.configure(width=canvas_width)

    def _get_all_rare_items_data_jjul(self):
        items_list_for_excel = []
        total_rare_item_value = 0
        all_rare_item_widgets = self.rare_items_manager_jjul.get_all_entry_widgets()
        msg_parent = self.rare_items_manager_jjul.get_messagebox_parent()

        for item_widgets_set in all_rare_item_widgets:
            try:
                name_widget = item_widgets_set['name_entry']
                value_widget = item_widgets_set['value_entry']
                item_name = MainApplication.static_get_entry_value(name_widget, str, "", field_name_for_error="고가템명", parent_for_msgbox=msg_parent)
                if not item_name: continue
                item_value = MainApplication.static_get_entry_value(value_widget, int, "0", field_name_for_error=f"'{item_name}' 가치", parent_for_msgbox=msg_parent)
                if item_name and item_value > 0:
                    items_list_for_excel.append((item_name, item_value))
                    total_rare_item_value += item_value
                elif item_name and item_value <= 0:
                    messagebox.showwarning("고가템 가치 오류", f"고가 아이템 '{item_name}'의 가치가 0 이하입니다. 계산에서 제외됩니다.", parent=msg_parent, detail="아이템 가치는 0보다 커야 합니다.")
            except ValueError: continue
        return items_list_for_excel, total_rare_item_value

    def _process_consumable_items_jjul(self):
        total_consumable_cost = 0
        total_consumable_gained_profit = 0
        all_consumable_widgets = self.consumables_manager_jjul.get_all_entry_widgets()
        msg_parent = self.consumables_manager_jjul.get_messagebox_parent()

        for item_widgets_set in all_consumable_widgets:
            try:
                name_widget = item_widgets_set['name_entry']
                price_value_widget = item_widgets_set['price_value_entry']
                start_qty_widget = item_widgets_set['start_qty_entry']
                end_qty_widget = item_widgets_set['end_qty_entry']

                item_name = MainApplication.static_get_entry_value(name_widget, str, "", field_name_for_error="소모/획득템 명칭", parent_for_msgbox=msg_parent)
                if not item_name: continue

                item_price_or_value = MainApplication.static_get_entry_value(price_value_widget, int, "0", field_name_for_error=f"'{item_name}' 개당 가격/가치", parent_for_msgbox=msg_parent)
                start_qty = MainApplication.static_get_entry_value(start_qty_widget, int, "0", field_name_for_error=f"'{item_name}' 시작 개수", parent_for_msgbox=msg_parent)
                end_qty = MainApplication.static_get_entry_value(end_qty_widget, int, "0", field_name_for_error=f"'{item_name}' 종료 개수", parent_for_msgbox=msg_parent)
                
                quantity_change = end_qty - start_qty

                if quantity_change < 0: 
                    used_quantity = -quantity_change
                    item_cost = used_quantity * item_price_or_value
                    total_consumable_cost += item_cost
                elif quantity_change > 0: 
                    gained_quantity = quantity_change
                    item_gained_profit = gained_quantity * item_price_or_value
                    total_consumable_gained_profit += item_gained_profit
            except ValueError: continue
        return total_consumable_cost, total_consumable_gained_profit

    def calculate_jjul_data(self, data_dict):
        calculated_data = data_dict.copy()
        try:
            start_dt_str = data_dict.get('시작시간'); end_dt_str = data_dict.get('종료시간')
            if not start_dt_str : messagebox.showerror("계산 오류", "쩔 시작시간이 입력되지 않았습니다.", parent=self.parent_frame); raise ValueError("쩔 시작시간 누락")
            if not end_dt_str : messagebox.showerror("계산 오류", "쩔 종료시간이 입력되지 않았습니다.", parent=self.parent_frame); raise ValueError("쩔 종료시간 누락")
            
            datetime.datetime.strptime(start_dt_str, "%H:%M")
            datetime.datetime.strptime(end_dt_str, "%H:%M")
        except ValueError as e:
            if "누락" not in str(e) and "잘못된 시간 형식" not in str(e) and "HH:MM" not in str(e): 
                messagebox.showerror("시간 형식 오류", "시작시간 또는 종료시간 형식이 잘못되었습니다 (HH:MM).", parent=self.parent_frame)
            raise

        calculated_data['총쩔비'] = data_dict.get('쩔인원수',0) * data_dict.get('1인당쩔비',0)
        
        sold_meso = data_dict.get('판매후메소', data_dict.get('종료메소',0))
        end_meso = data_dict.get('종료메소',0)
        if sold_meso >= end_meso:
            calculated_data['일반템수익'] = sold_meso - end_meso
        else:
            messagebox.showwarning("메소 입력 오류 (쩔)", "판매 후 메소가 종료 메소보다 적습니다. 일반템 수익은 0으로 처리됩니다.", parent=self.parent_frame)
            calculated_data['일반템수익'] = 0
            
        rare_items_list, total_rare_value = self._get_all_rare_items_data_jjul()
        calculated_data['rare_items_list'] = rare_items_list
        calculated_data['고가아이템가치'] = total_rare_value
        calculated_data['고가아이템'] = ", ".join([item[0] for item in rare_items_list]) if rare_items_list else ""
        
        total_consumable_cost, total_consumable_gained_profit = self._process_consumable_items_jjul()
        calculated_data['소모아이템비'] = total_consumable_cost
        calculated_data['소모템획득수익'] = total_consumable_gained_profit
        
        calculated_data['총수익'] = (calculated_data.get('총쩔비',0) + 
                                   calculated_data.get('일반템수익',0) + 
                                   total_rare_value +
                                   total_consumable_gained_profit)
        calculated_data['순수익'] = calculated_data.get('총수익',0) - calculated_data.get('소모아이템비',0)
        
        return calculated_data

    def save_data(self):
        raw_data_from_form = {}
        try:
            excel_file_path = self.main_app.excel_manager.filename
            if not os.path.exists(excel_file_path):
                if not self.main_app.excel_manager.create_excel_template():
                    messagebox.showerror("템플릿 생성 실패", f"'{os.path.basename(excel_file_path)}' 템플릿 생성 실패.", parent=self.parent_frame)
                    return
                else:
                    messagebox.showinfo("템플릿 자동 생성", f"'{os.path.basename(excel_file_path)}' 새로 생성.", parent=self.parent_frame)

            raw_data_from_form['날짜'] = MainApplication.static_get_entry_value(self.date_entry, str, field_name_for_error="날짜", parent_for_msgbox=self.parent_frame)
            if not raw_data_from_form['날짜']: messagebox.showerror("입력 오류", "날짜를 입력해주세요.", parent=self.parent_frame); self.date_entry.focus_set(); return
            
            raw_data_from_form['맵명'] = MainApplication.static_get_entry_value(self.map_name_entry, str, field_name_for_error="맵명", parent_for_msgbox=self.parent_frame)
            if not raw_data_from_form['맵명']: messagebox.showerror("입력 오류", "맵명을 입력해주세요.", parent=self.parent_frame); self.map_name_entry.focus_set(); return
            
            raw_data_from_form['시작시간'] = self.start_time_entry.get().strip()
            raw_data_from_form['종료시간'] = self.end_time_entry.get().strip()
            if not raw_data_from_form['시작시간']: messagebox.showerror("입력 오류", "시작시간을 입력해주세요.", parent=self.parent_frame); self.start_time_entry.focus_set(); return
            if not raw_data_from_form['종료시간']: messagebox.showerror("입력 오류", "종료시간을 입력해주세요.", parent=self.parent_frame); self.end_time_entry.focus_set(); return
            try: 
                datetime.datetime.strptime(raw_data_from_form['시작시간'], "%H:%M")
                datetime.datetime.strptime(raw_data_from_form['종료시간'], "%H:%M")
            except ValueError:
                messagebox.showerror("입력 오류", "시작시간 또는 종료시간 형식이 올바르지 않습니다 (HH:MM).", parent=self.parent_frame); return

            raw_data_from_form['시작메소'] = MainApplication.static_get_entry_value(self.start_meso_entry, int, "0", field_name_for_error="시작 메소", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['종료메소'] = MainApplication.static_get_entry_value(self.end_meso_entry, int, "0", field_name_for_error="종료 메소", parent_for_msgbox=self.parent_frame)
            
            sold_meso_str = self.sold_meso_entry.get().strip()
            if not sold_meso_str or sold_meso_str == "0": raw_data_from_form['판매후메소'] = raw_data_from_form['종료메소']
            else: raw_data_from_form['판매후메소'] = MainApplication.static_get_entry_value(self.sold_meso_entry, int, str(raw_data_from_form['종료메소']), field_name_for_error="판매 후 메소", parent_for_msgbox=self.parent_frame)
            
            raw_data_from_form['쩔인원수'] = MainApplication.static_get_entry_value(self.party_size_entry, int, "0", field_name_for_error="쩔 인원 수", parent_for_msgbox=self.parent_frame)
            raw_data_from_form['1인당쩔비'] = MainApplication.static_get_entry_value(self.price_per_person_entry, int, "0", field_name_for_error="1인당 쩔비", parent_for_msgbox=self.parent_frame)
            
            final_data_to_save = self.calculate_jjul_data(raw_data_from_form)
            
            if self.main_app.excel_manager.save_data_to_sheet(final_data_to_save, "쩔세션", JJUL_SHEET_COLUMNS, "쩔", parent_for_msgbox=self.parent_frame):
                date_to_update = final_data_to_save.get('날짜')
                jjul_profit_val = final_data_to_save.get('총쩔비', 0) 
                rare_item_val = final_data_to_save.get('고가아이템가치', 0)
                normal_item_val = final_data_to_save.get('일반템수익', 0)
                consumable_gained_val = final_data_to_save.get('소모템획득수익', 0)
                consumable_cost_val = final_data_to_save.get('소모아이템비', 0)

                self.main_app.excel_manager.update_daily_summary(
                    date_str=date_to_update,
                    jjul_profit=jjul_profit_val,
                    rare_item_profit=rare_item_val,
                    normal_item_profit=normal_item_val,
                    consumable_gained_profit=consumable_gained_val,
                    consumable_cost=consumable_cost_val,
                    parent_for_msgbox=self.parent_frame
                )
                map_name_to_update_jjul = final_data_to_save.get('맵명')
                session_main_profit_jjul = final_data_to_save.get('총쩔비', 0) + final_data_to_save.get('일반템수익', 0)
                session_rare_profit_jjul = final_data_to_save.get('고가아이템가치', 0)
                session_consumable_gained_jjul = final_data_to_save.get('소모템획득수익', 0)

                self.main_app.excel_manager.update_map_summary(
                    map_name=map_name_to_update_jjul,
                    session_type="쩔",
                    profit_from_session=session_main_profit_jjul,
                    rare_item_profit_from_session=session_rare_profit_jjul,
                    consumable_gained_profit_from_session=session_consumable_gained_jjul,
                    parent_for_msgbox=self.parent_frame
                )
            
                date_for_weekday_update_jjul = final_data_to_save.get('날짜')
                session_type_jjul = "쩔"
                val_pure_revenue_jjul = final_data_to_save.get('총쩔비', 0)
                val_rare_item_jjul = final_data_to_save.get('고가아이템가치', 0)
                val_normal_item_jjul = final_data_to_save.get('일반템수익', 0)
                val_consumable_gained_jjul = final_data_to_save.get('소모템획득수익', 0)
                val_net_profit_jjul = final_data_to_save.get('순수익', 0)

                self.main_app.excel_manager.update_weekday_summary(
                    date_str=date_for_weekday_update_jjul,
                    session_type=session_type_jjul,
                    val_pure_revenue_contribution=val_pure_revenue_jjul,
                    val_rare_item_contribution=val_rare_item_jjul,
                    val_normal_item_contribution=val_normal_item_jjul,
                    val_consumable_gained_contribution=val_consumable_gained_jjul,
                    val_net_profit_contribution=val_net_profit_jjul,
                    parent_for_msgbox=self.parent_frame
                )
                current_filename = os.path.basename(self.main_app.excel_manager.filename)
                messagebox.showinfo("저장 완료", f"쩔 결과가 '{current_filename}'에 저장되고 관련 통계가 업데이트되었습니다.", parent=self.parent_frame)
                self.clear_fields()
        except ValueError as e: 
            print(f"데이터 입력/계산 오류로 저장 중단 (JjulResultForm): {e}"); 
            return
        except Exception as e: 
            messagebox.showerror("알 수 없는 오류", f"쩔 데이터 저장 중 예기치 않은 오류 발생: {e}", parent=self.parent_frame); 
            print(f"알 수 없는 오류 (JjulResultForm.save_data): {e}"); 
            import traceback; traceback.print_exc(); 
            return

    def clear_fields(self):
        self.date_entry.delete(0, tk.END); self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.map_name_entry.delete(0, tk.END)
        self.start_time_entry.delete(0, tk.END); self.end_time_entry.delete(0, tk.END)
        self.start_meso_entry.delete(0, tk.END); self.start_meso_entry.insert(0, "0")
        self.end_meso_entry.delete(0, tk.END); self.end_meso_entry.insert(0, "0")
        self.sold_meso_entry.delete(0, tk.END); self.sold_meso_entry.insert(0, "0")
        self.party_size_entry.delete(0, tk.END); self.party_size_entry.insert(0, "0")
        self.price_per_person_entry.delete(0, tk.END); self.price_per_person_entry.insert(0, "0")
        
        self.rare_items_manager_jjul.clear_all_item_entries()
        self.consumables_manager_jjul.clear_all_item_entries()
        
        self.map_name_entry.focus_set(); print("쩔 결과 필드 초기화됨.")
        self.canvas.after_idle(lambda: (
            self.canvas.configure(scrollregion=self.canvas.bbox("all")),
            self.canvas.yview_moveto(0)
        ))


class MesoSaleForm:
    def __init__(self, parent_frame, main_app_ref):
        self.parent_frame = parent_frame
        self.main_app = main_app_ref
        self.main_content_frame = ttk.Frame(parent_frame, padding="20")
        self.main_content_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        current_row = 0

        lbl_ms_date = ttk.Label(self.main_content_frame, text="판매 날짜:")
        lbl_ms_date.grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.date_entry = ttk.Entry(self.main_content_frame, width=30)
        self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.date_entry.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        date_tooltip_text = "날짜 (YYYY-MM-DD). 'MMDD', 'YYMMDD' 등 입력 시 자동 완성."
        Tooltip(self.date_entry, date_tooltip_text); Tooltip(lbl_ms_date, date_tooltip_text)
        self.date_entry.bind("<FocusOut>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))
        self.date_entry.bind("<Return>", lambda event, entry=self.date_entry: MainApplication.static_format_date_entry(event, entry, self.parent_frame))
        current_row += 1

        lbl_ms_total = ttk.Label(self.main_content_frame, text="총 판매 메소량:")
        lbl_ms_total.grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.total_meso_sold_entry = ttk.Entry(self.main_content_frame, width=30); self.total_meso_sold_entry.insert(0, "0")
        self.total_meso_sold_entry.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        Tooltip(self.total_meso_sold_entry, "판매한 총 메소량 (예: 1.5억, 2300만, 150000000)")
        Tooltip(lbl_ms_total, "판매한 총 메소량 (예: 1.5억, 2300만, 150000000)")
        
        lbl_ms_example = ttk.Label(self.main_content_frame, text="(예: 1.5억, 2300만, 150000000)")
        lbl_ms_example.grid(row=current_row, column=2, padx=5, pady=8, sticky=tk.W)
        Tooltip(lbl_ms_example, "입력 예시입니다.")
        current_row += 1

        lbl_ms_price = ttk.Label(self.main_content_frame, text="100만 메소당 단가(원):")
        lbl_ms_price.grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.price_per_1m_entry = ttk.Entry(self.main_content_frame, width=30); self.price_per_1m_entry.insert(0, "0")
        self.price_per_1m_entry.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        Tooltip(self.price_per_1m_entry, "100만 메소당 판매 가격 (원 단위)")
        Tooltip(lbl_ms_price, "100만 메소당 판매 가격 (원 단위)")
        current_row += 1

        lbl_ms_value = ttk.Label(self.main_content_frame, text="총 판매액(원):")
        lbl_ms_value.grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.total_sale_value_label = ttk.Label(self.main_content_frame, text="0 원", width=30, anchor=tk.W)
        self.total_sale_value_label.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        Tooltip(self.total_sale_value_label, "자동으로 계산된 총 판매 금액입니다.")
        Tooltip(lbl_ms_value, "자동으로 계산된 총 판매 금액입니다.")
        current_row += 1

        self.total_meso_sold_entry.bind("<KeyRelease>", self.update_total_sale_value)
        self.price_per_1m_entry.bind("<KeyRelease>", self.update_total_sale_value)
        
        self.action_buttons_frame = ttk.Frame(self.main_content_frame)
        self.action_buttons_frame.grid(row=current_row, column=0, columnspan=3, pady=20)
        self.save_button = ttk.Button(self.action_buttons_frame, text="판매 기록 저장", command=self.save_data)
        self.save_button.pack(side=tk.LEFT, padx=10); Tooltip(self.save_button, "입력한 메소 판매 기록을 저장합니다.")
        self.clear_button = ttk.Button(self.action_buttons_frame, text="초기화", command=self.clear_fields)
        self.clear_button.pack(side=tk.LEFT, padx=10); Tooltip(self.clear_button, "모든 입력 필드를 초기화합니다.")
        
        self.main_content_frame.columnconfigure(1, weight=1)
        self.main_content_frame.columnconfigure(2, weight=0)

    def _parse_meso_string(self, meso_str_input):
        meso_str = meso_str_input.strip().replace(",", "").lower()
        total_mesos = 0
        if not meso_str: return 0

        억_pattern = r'([\d\.]+)\s*억'
        억_match = re.search(억_pattern, meso_str)
        if 억_match:
            try:
                total_mesos += float(억_match.group(1).rstrip('.')) * 100_000_000
            except ValueError: 
                print(f"경고: '억' 단위 파싱 중 잘못된 숫자 형식: '{억_match.group(1)}'")
            meso_str = re.sub(억_pattern, "", meso_str, 1).strip()

        만_pattern = r'([\d\.]+)\s*만'
        만_match = re.search(만_pattern, meso_str)
        if 만_match:
            try:
                total_mesos += float(만_match.group(1).rstrip('.')) * 10_000
            except ValueError:
                print(f"경고: '만' 단위 파싱 중 잘못된 숫자 형식: '{만_match.group(1)}'")
            meso_str = re.sub(만_pattern, "", meso_str, 1).strip()

        if meso_str:
            try:
                total_mesos += float(meso_str.rstrip('.'))
            except ValueError:
                print(f"경고: 메소 문자열 파싱 중 숫자 변환 불가 부분 발견: '{meso_str}'")
        
        return int(total_mesos)

    def update_total_sale_value(self, event=None):
        try:
            total_meso_sold_str = self.total_meso_sold_entry.get()
            price_per_1m_str = self.price_per_1m_entry.get().replace(",", "")

            total_mesos_numerical = self._parse_meso_string(total_meso_sold_str)
            if total_mesos_numerical < 0: total_mesos_numerical = 0

            sale_amount_1m_unit = total_mesos_numerical / 1_000_000
            
            price_per_1m = 0
            if not price_per_1m_str:
                price_per_1m = 0
            elif price_per_1m_str.lstrip('-').isdigit():
                price_per_1m = int(price_per_1m_str)
            else:
                self.total_sale_value_label.config(text="단가 오류")
                return
            
            total_value = int(price_per_1m * sale_amount_1m_unit)
            self.total_sale_value_label.config(text=f"{total_value:,} 원")
        except ValueError:
            self.total_sale_value_label.config(text="계산 오류")
        except Exception as e:
            print(f"MesoSaleForm.update_total_sale_value 오류: {e}")
            self.total_sale_value_label.config(text="입력값 오류")


    def save_data(self):
        raw_data_for_excel = {}
        try:
            excel_file_path = self.main_app.excel_manager.filename
            if not os.path.exists(excel_file_path):
                if not self.main_app.excel_manager.create_excel_template():
                    messagebox.showerror("템플릿 생성 실패", f"'{os.path.basename(excel_file_path)}' 템플릿 생성 실패.", parent=self.parent_frame)
                    return
                else:
                    messagebox.showinfo("템플릿 자동 생성", f"'{os.path.basename(excel_file_path)}' 새로 생성.", parent=self.parent_frame)

            raw_data_for_excel['날짜'] = MainApplication.static_get_entry_value(self.date_entry, str, field_name_for_error="판매 날짜", parent_for_msgbox=self.parent_frame)
            if not raw_data_for_excel['날짜']: messagebox.showwarning("입력 오류", "판매 날짜를 입력해주세요.", parent=self.parent_frame); self.date_entry.focus_set(); return

            total_meso_sold_str = self.total_meso_sold_entry.get()
            total_mesos_numerical = self._parse_meso_string(total_meso_sold_str)
            
            if total_mesos_numerical <= 0: 
                messagebox.showwarning("입력 오류", "총 판매 메소량은 0보다 커야 합니다.", parent=self.parent_frame)
                self.total_meso_sold_entry.focus_set(); return
            raw_data_for_excel['판매량(단위: 100만메소)'] = total_mesos_numerical / 1_000_000

            price_per_1m_val = MainApplication.static_get_entry_value(self.price_per_1m_entry, int, "0", field_name_for_error="100만 메소당 단가", parent_for_msgbox=self.parent_frame)
            if price_per_1m_val <= 0: 
                messagebox.showwarning("입력 오류", "100만 메소당 단가는 0보다 커야 합니다.", parent=self.parent_frame)
                self.price_per_1m_entry.focus_set(); return
            raw_data_for_excel['100만메소당가격(원)'] = price_per_1m_val
            
            raw_data_for_excel['총판매액(원)'] = int(raw_data_for_excel['100만메소당가격(원)'] * raw_data_for_excel['판매량(단위: 100만메소)'])
            
            if self.main_app.excel_manager.save_data_to_sheet(raw_data_for_excel, "메소판매기록", MESO_SALE_SHEET_COLUMNS, parent_for_msgbox=self.parent_frame):
                date_to_update = raw_data_for_excel.get('날짜')
                cash_sold_val = raw_data_for_excel.get('총판매액(원)', 0)

                self.main_app.excel_manager.update_daily_summary(
                    date_str=date_to_update,
                    cash_sold_krw=cash_sold_val,
                    parent_for_msgbox=self.parent_frame
                )
                current_filename = os.path.basename(self.main_app.excel_manager.filename)
                messagebox.showinfo("저장 완료", f"메소 판매 기록이 '{current_filename}'에 저장되고 일별 요약이 업데이트되었습니다.", parent=self.parent_frame)
                self.clear_fields()
        except ValueError as ve:
            print(f"메소 판매 데이터 처리 중단 (ValueError): {ve}")
            return
        except Exception as e:
            messagebox.showerror("알 수 없는 오류", f"메소 판매 데이터 처리 중 예기치 않은 오류 발생: {e}", parent=self.parent_frame)
            print(f"알 수 없는 오류 (MesoSaleForm.save_data): {e}")
            import traceback; traceback.print_exc()
            return

    def clear_fields(self):
        self.date_entry.delete(0, tk.END); self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.total_meso_sold_entry.delete(0, tk.END); self.total_meso_sold_entry.insert(0, "0")
        self.price_per_1m_entry.delete(0, tk.END); self.price_per_1m_entry.insert(0, "0")
        self.total_sale_value_label.config(text="0 원")
        self.total_meso_sold_entry.focus_set()
        print("메소 판매 필드 초기화됨.")

if __name__ == "__main__":
    root = tk.Tk()
    app = MainApplication(root)

    def set_initial_focus():
        target_widget = None
        if hasattr(app, 'hunting_form') and hasattr(app.hunting_form, 'map_name_entry'):
            target_widget = app.hunting_form.map_name_entry
        
        if target_widget:
            target_widget.focus_set()
        elif hasattr(app, 'root'): 
            app.root.focus_force()

    root.after(100, set_initial_focus) 
    root.mainloop()
